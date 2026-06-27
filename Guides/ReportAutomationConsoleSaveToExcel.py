from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from datetime import datetime
from selenium.webdriver.common.action_chains import ActionChains
import os
import glob
import pandas as pd
import time
import openpyxl
import arabic_reshaper
import math
import re
import win32com.client
from selenium.common.exceptions import TimeoutException
from bidi.algorithm import get_display
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import os
from openpyxl import Workbook, load_workbook
import json
from cryptography.fernet import Fernet
from datetime import datetime, timedelta
from selenium.common.exceptions import StaleElementReferenceException, ElementClickInterceptedException, WebDriverException
from openpyxl.styles import Font, Alignment, Border, Side ,PatternFill
from openpyxl.utils import get_column_letter
import warnings


# ==========================
#region Functions Section
# ==========================

def select_second_option_from_dropdown(dropdown_button_xpath, expected_option_text):
    try:
        # Click the dropdown to open
        dropdown = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, dropdown_button_xpath)))
        print_arabic(f"📂 تم العثور على زر القائمة المنسدلة")     

        driver.execute_script("arguments[0].click();", dropdown)
        print_arabic(f"📂 تم النقر على زر القائمة المنسدلة")

        # Wait for the dropdown overlay to appear and contain options
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, "//div[contains(@class,'dx-overlay-content') and contains(@style,'visibility: visible')]"))
        )
        
        # Wait explicitly for overlay to refresh and contain the expected option (by text)
        second_option = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, f"//div[contains(@class,'dx-overlay-content') and contains(@style,'visibility: visible')]//div[contains(@class,'dx-item') and contains(., '{expected_option_text}')]"))
        )

        # Scroll into view and click using ActionChains to avoid interception
        driver.execute_script("arguments[0].scrollIntoView(true);", second_option)
        ActionChains(driver).move_to_element(second_option).click().perform()
        print_arabic(f"📂 تم اختيار الخيار الثاني من القائمة")

    except Exception as e:
        print_arabic(f"❌ خطأ أثناء الاختيار من القائمة المنسدلة: {e}")

def wait_for_download_complete(folder_path, timeout=60):
    for _ in range(timeout):
        cr_files = glob.glob(os.path.join(folder_path, "*.crdownload"))
        xlsx_files = glob.glob(os.path.join(folder_path, "*.xlsx"))
        if not cr_files and xlsx_files:
            return max(xlsx_files, key=os.path.getctime)
        time.sleep(1)
    return None

def clear_folder(folder_path):
    files = glob.glob(os.path.join(folder_path, "*"))
    for file_path in files:
        try:
            os.remove(file_path)
        except Exception as e:
            print_arabic(f"⚠️ لم يتم حذف الملف: {file_path} بسبب الخطأ: {e}")

def export_report_to_excel_fixed(folder_path, name=None):
    try:
        global driver

        original_tab = driver.current_window_handle
        initial_tabs = driver.window_handles

        WebDriverWait(driver, 10).until(lambda d: len(d.window_handles) > len(initial_tabs))
        new_tab = [tab for tab in driver.window_handles if tab != original_tab][0]
        driver.switch_to.window(new_tab)
        print_arabic("🔄 تم الانتقال إلى صفحة تحميل التقرير")
        time.sleep(10)
        print_arabic("📄 يتم تحميل بيانات التقرير, الرجاء الانتظار ...")

        WebDriverWait(driver, 300).until(
            EC.invisibility_of_element_located((By.ID, "repViewer_AsyncWait_Wait"))
        )
        print_arabic("🔄 تم الانتهاء من عميلة تحميل التقرير")

        WebDriverWait(driver, 60).until(
            EC.visibility_of_element_located((By.XPATH, "//div[@id='VisibleReportContentrepViewer_ctl13']"))
        )
        print_arabic("✅ التقرير مرئي")

        export_button = WebDriverWait(driver, 60).until(
            EC.element_to_be_clickable((By.ID, "repViewer_ctl09_ctl04_ctl00_ButtonLink"))
        )
        driver.execute_script("arguments[0].click();", export_button)
        print_arabic("✅ تم النقر على قائمة تصدير إلى ")

        excel_option = WebDriverWait(driver, 60).until(
            EC.element_to_be_clickable((By.XPATH, "//a[@title='Excel']"))
        )
        before = set(glob.glob(os.path.join(folder_path, "*.xlsx")))
        driver.execute_script("arguments[0].click();", excel_option)
        print_arabic("✅ التصدير إلى ملف أكسل")

        downloaded_file = None
        for _ in range(60):
            after = set(glob.glob(os.path.join(folder_path, "*.xlsx")))
            new_files = after - before
            partial_files = glob.glob(os.path.join(folder_path, "*.crdownload"))
            if new_files and not partial_files:
                downloaded_file = list(new_files)[0]
                break
            time.sleep(1)

        if not downloaded_file:
            print_arabic("❌ لم يتم العثور على ملف Excel تم تنزيله")
            return

        if name:
            base_name = f"{name}.xlsx" if not name.endswith(".xlsx") else name
            new_path = os.path.join(folder_path, base_name)
            os.rename(downloaded_file, new_path)
            print_arabic(f"✅ تم حفظ الملف باسم: {base_name}")
        else:
            print_arabic(f"✅ تم تحميل الملف: {os.path.basename(downloaded_file)}")

        driver.close()
        driver.switch_to.window(original_tab)
        print_arabic("🔙 الانتقال إلى صفحة الويب السابقة")

    except Exception as e:
        print_arabic(f"❌ خطأ أثناء عملية تحميل الملف: {e}")

def get_latest_excel_file(download_folder, filename_startswith):
    # Build search pattern (to cover file with and without (1), (2), etc.)
    search_pattern = os.path.join(download_folder, f"{filename_startswith}*.xlsx")
    
    # Find all matching files
    matching_files = glob.glob(search_pattern)
    if not matching_files:
        print_arabic(f"🚫 لا يوجد ملفات مطابقة")
        return None

    # Sort by modification time (latest first)
    latest_file = max(matching_files, key=os.path.getmtime)
    # print_arabic(f"✅ أحدث ملف موجود هو باسم: {latest_file}")
    return latest_file

def compare_values(label1, value1, label2, value2,Param = 0):
    try:
        value1 = float(value1)
        value2 = float(value2)
        if Param == 0:
            if value1 > value2:
                bigger, smaller = label1, label2
                print_arabic(f"{bigger} أكبر من {smaller}")
                print_arabic(f"الفرق هو: {abs(value1 - value2)}")
            elif value2 > value1:
                bigger, smaller = label2, label1
                print_arabic(f"{bigger} أكبر من {smaller}")
                print_arabic(f"الفرق هو: {abs(value2 - value1)}")
            else:
                print_arabic(f"{label1} و {label2} متساويين.")
        else :
            if value1 > value2:
                bigger, smaller = label1, label2
                print_arabic(f"{bigger} أكبر من {smaller}")
                print_arabic(f"الفرق هو: {abs(value1 + value2)}")
            elif value2 > value1:
                bigger, smaller = label2, label1
                print_arabic(f"{bigger} أكبر من {smaller}")
                print_arabic(f"الفرق هو: {abs(value2 + value1)}")
            else:
                print_arabic(f"{label1} و {label2} متساويين")
    except Exception as e:
         print_arabic(f"❌ خطأ في المقارنة بين القيم: {e}")

def column_letter_to_index(columns):
    if isinstance(columns, str):  # Single column letter
        return openpyxl.utils.column_index_from_string(columns)
    elif isinstance(columns, list):  # Multiple column letters
        return [openpyxl.utils.column_index_from_string(col) for col in columns]
    return columns  # If columns are already indices (integers)

def get_last_value_in_column(file_path, sheet_name, columns):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    # Handle single vs multiple columns
    is_single_column = isinstance(columns, str)
    col_indices = column_letter_to_index(columns)

    if is_single_column:
        col_index = col_indices
        for row in range(ws.max_row, 0, -1):
            cell_value = ws.cell(row=row, column=col_index).value
            if cell_value is not None and str(cell_value).strip() != "":
                return cell_value
        print(f"⚠ No non-empty values found in column {columns}.")
        return None
    else:
        # Check merged ranges that include any of the specified columns
        for row in range(ws.max_row, 0, -1):
            for merged_range in ws.merged_cells.ranges:
                if merged_range.min_row <= row <= merged_range.max_row:
                    if any(merged_range.min_col <= col <= merged_range.max_col for col in col_indices):
                        value = ws.cell(merged_range.min_row, merged_range.min_col).value
                        if value is not None and str(value).strip() != "":
                            return value
        print(f"⚠ No merged values found in columns {columns}.")
        return None

def print_arabic(text):
    reshaped_text = arabic_reshaper.reshape(text)
    bidi_text = get_display(reshaped_text)
    print(bidi_text)

    # Save to log file in the created folder
    try:
        if 'folder_path' in globals():  # Only log if folder is defined
            with open(os.path.join(folder_path, "log.txt"), "a", encoding="utf-8") as log_file:
                log_file.write(text + "\n")
    except Exception as e:
        print(f"[LOG ERROR] Could not write to log file: {e}")

def get_valueOf_receipt_document(file_path, label):
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active

        # نبحث عن الصف الذي يحتوي على "سند قبض" في الأعمدة K إلى O (11 إلى 15)
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=11, max_col=15):
            for cell in row:
                if cell.value and str(cell.value).strip() == label:
                    receipt_row = cell.row
                    # نأخذ القيمة من العمود G في نفس الصف
                    value = sheet.cell(row=receipt_row, column=7).value  # G = 7
                    return value

        print(f"❌ لم يتم العثور على '{label}' في الأعمدة K إلى O")
        return None

    except Exception as e:
        print(f"❌ خطأ أثناء المعالجة: {e}")
        return None

# def SaveToExcel(label, value):
#     filename = "output.xlsx"
#     full_path = os.path.join(folder_path, filename)

#     # Load or create the workbook
#     if os.path.exists(full_path):
#         wb = load_workbook(full_path)
#         ws = wb.active
#     else:
#         wb = Workbook()
#         ws = wb.active

#     # Find the next empty row
#     next_row = ws.max_row + 1 if ws.max_row > 0 else 1

#     # Write label and value
#     ws.cell(row=next_row, column=1, value=label)
#     ws.cell(row=next_row, column=2, value=value)

#     # Save the workbook
#     wb.save(full_path)

def SaveToExcel(label, value, folder_path, filename="output.xlsx", keep_full_path=False):
    full_path = os.path.join(folder_path, filename)
    header_names = ["Label", "Value"]

    # Ensure folder exists
    os.makedirs(folder_path, exist_ok=True)

    # Load or create workbook/sheet
    if os.path.exists(full_path):
        wb = load_workbook(full_path)
        ws = wb.active

        # If headers not present in first row, insert them (push existing rows down)
        first_col = (ws.cell(row=1, column=1).value or "")
        second_col = (ws.cell(row=1, column=2).value or "")
        if not (str(first_col).strip().lower() == header_names[0].lower()
                and str(second_col).strip().lower() == header_names[1].lower()):
            ws.insert_rows(1)
            ws.cell(row=1, column=1, value=header_names[0])
            ws.cell(row=1, column=2, value=header_names[1])

            # style header row after insertion
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            for col in (1, 2):
                c = ws.cell(row=1, column=col)
                c.font = Font(bold=True, color="FFFFFF")
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.fill = header_fill
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(header_names)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for col in (1, 2):
            c = ws.cell(row=1, column=col)
            c.font = Font(bold=True, color="FFFFFF")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.fill = header_fill

    # Prepare value: if it's a path to an existing file and keep_full_path is False, use basename
    value_to_write = value
    try:
        if (isinstance(value, str)
                and not keep_full_path
                and os.path.exists(value)
                and os.path.isfile(value)):
            value_to_write = os.path.basename(value)
    except Exception:
        # On any error, fall back to str(value)
        value_to_write = value

    # Convert non-strings safely
    if not isinstance(value_to_write, (str, int, float)):
        try:
            value_to_write = str(value_to_write)
        except Exception:
            value_to_write = repr(value_to_write)

    # Write to next row
    next_row = ws.max_row + 1
    ws.cell(row=next_row, column=1, value=label)
    ws.cell(row=next_row, column=2, value=value_to_write)

    # Apply styling/borders for the new row
    thin = Side(style='thin')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in (1, 2):
        cell = ws.cell(row=next_row, column=col)
        cell.border = thin_border
        # label column: wrap text for long Arabic labels
        if col == 1:
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        else:
            cell.alignment = Alignment(vertical="center")

    # Auto-adjust column widths (simple heuristic)
    for col_idx in (1, 2):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for row in range(1, ws.max_row + 1):
            cell = ws[f"{col_letter}{row}"]
            if cell.value is None:
                continue
            length = len(str(cell.value))
            if length > max_length:
                max_length = length
        # minimum width and small padding
        ws.column_dimensions[col_letter].width = max(12, max_length + 2)

    # Save workbook
    wb.save(full_path)
    print_arabic(f"✅ Saved to {full_path} -> {label} = {value_to_write}")
    return full_path

def select_payment_methods(driver, methods_to_select, dropdown_xpath, ok_button_xpath):
    # dropdown_xpath = '//*[@id="kt_content"]/div/app-show/div/div[1]/div[3]/div[2]/dx-tag-box/div[1]/div/div[1]'
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, dropdown_xpath))).click()

    elements = driver.find_elements(By.XPATH, "//div[contains(@class,'dx-list-item')]")

    for method in methods_to_select:
        found = False
        for el in elements:
            # Get full visible text of the div
            el_text = driver.execute_script("return arguments[0].innerText;", el).strip()
            if method in el_text:
                driver.execute_script("arguments[0].scrollIntoView(true);", el)
                el.click()
                found = True
                break
        if not found:
            print(f"⚠️ Could not find/select: {method}")

    # Click موافق button
    # try:
    #     ok_button = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div/div[2]/div/div[2]/div[1]/div/div"))) 
    #     driver.execute_script("arguments[0].scrollIntoView(true);", ok_button)
    #     ok_button.click()    
    # except Exception as e: 
    #     print(f"⚠️ Could not click موافق button: {e}")

    #     try:
    #         ok_button = driver.find_element(By.XPATH, "/html/body/div[2]/div/div[2]/div/div[2]/div[1]/div/div")
    #         driver.execute_script("arguments[0].click();", ok_button)
    #     except Exception as e2:
    #         print(f"⚠️ JS click also failed: {e2}")
    try:
        ok_button = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.XPATH, ok_button_xpath))
        )
        driver.execute_script("arguments[0].scrollIntoView(true);", ok_button)
        ok_button.click()
    except Exception as e:
        print(f"⚠️ Could not click موافق button: {e}")
        try:
            ok_button = driver.find_element(By.XPATH, ok_button_xpath)
            driver.execute_script("arguments[0].click();", ok_button)
        except Exception as e2:
            print(f"⚠️ JS click also failed: {e2}")

def TotalReportOfTransactionsbyRevenueSourceRec(driver, from_date_arabic, to_date_arabic, folder_path):
    report_url = "https://tahseel.gov.ae/ManagePortal/report-show/7c9f7dcd-1163-4e89-91dd-02b841c24ed7"
    driver.get(report_url)
    print_arabic("🌐 الانتقال إلى صفحة تقرير المقبوضات")
    time.sleep(5)
    
    select_second_option_from_dropdown(
        "//*[@id='kt_content']/div/app-show/div/div[1]/div[12]/div[2]/dx-select-box//div[contains(@class,'dx-dropdowneditor-button')]",
        "مدفوعة"
    )
    from_date_input = driver.find_element(By.XPATH, "//*[@id='kt_content']/div/app-show/div/div[1]/div[7]/div[2]/dx-date-box//input[@type='text']")
    to_date_input = driver.find_element(By.XPATH, "//*[@id='kt_content']/div/app-show/div/div[1]/div[8]/div[2]/dx-date-box//input[@type='text']")

    from_date_input.click()
    from_date_input.clear()
    time.sleep(0.5)
    from_date_input.send_keys(from_date_arabic)
    from_date_input.send_keys(Keys.TAB)

    # Clear and set TO date
    to_date_input.click()
    to_date_input.clear()
    time.sleep(0.5)
    to_date_input.send_keys(to_date_arabic)
    to_date_input.send_keys(Keys.TAB)

    print_arabic("✅ تم تحديد التواريخ بنجاح")

    original_tab = driver.current_window_handle
    initial_tabs = driver.window_handles

    submit_button = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, '//*[@id="kt_content"]/div/app-show/div/div[2]/button'))
    )
    driver.execute_script("arguments[0].click();", submit_button)
    print_arabic("✅ تم الضغط على زر عرض تقرير إجمالي المعاملات حسب جهة الإيراد - المقبوضات")

    export_report_to_excel_fixed(folder_path)

def TransactionsFeeReport(driver, from_date_arabic, to_date_arabic, folder_path):
    transactionFee_url2 = "https://tahseel.gov.ae/ManagePortal/report-show/97353334-399a-4613-9097-9cf5dc95c690"
    driver.get(transactionFee_url2)
    time.sleep(5)
    print_arabic(f"🌐 الانتقال إلى صفحة تقرير رسوم المعاملات لكافة وسائل الدفع")

    try:
        wait = WebDriverWait(driver, 15)

        # Locate the radio button by its text inside .dx-item-content, then go to its parent .dx-item
        radio_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//div[@class='dx-item-content' and contains(text(), 'معاملات ايراد')]/parent::div[@role='radio']")))
            
        # Confirm and click
        if radio_button.is_displayed() and radio_button.is_enabled():
            driver.execute_script("arguments[0].scrollIntoView();", radio_button)
            radio_button.click()
            print_arabic(f"✅  تم تحديد خيار معاملات إيراد في تقرير كافة وسائل الدفع بنجاح")
        else:
            print_arabic(f"❌  لم يتم اختيار خيار معاملات الإيراد في تقرير كافة وسائل الدفع")

    except Exception as e:
        print_arabic(f"حدثت المشكلة التالية في تقرير كافة وسائل الدفع:{e}")

    from_date_input2 = driver.find_element(By.XPATH, "//*[@id='kt_content']/div/app-show/div/div[1]/div[5]/div[2]/dx-date-box//input[@type='text']")
    to_date_input2 = driver.find_element(By.XPATH, "//*[@id='kt_content']/div/app-show/div/div[1]/div[6]/div[2]/dx-date-box//input[@type='text']")

    from_date_input2.click()
    from_date_input2.clear()
    time.sleep(0.5)
    from_date_input2.send_keys(from_date_arabic)
    from_date_input2.send_keys(Keys.TAB)

    to_date_input2.click()
    to_date_input2.clear()
    time.sleep(0.5)
    to_date_input2.send_keys(to_date_arabic)
    to_date_input2.send_keys(Keys.TAB)

    print_arabic(f"✅ تم ضبط تواريخ البحث لتقرير كافة وسائل الدفع")

    submit_button2 = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="kt_content"]/div/app-show/div/div[2]/button')))
    driver.execute_script("arguments[0].click();", submit_button2)
    print_arabic(f"✅ تم الضغط على زر عرض تقرير  كافة وسائل الدفع")

    export_report_to_excel_fixed(folder_path)

def UniversalPayments(driver, from_date_arabic, to_date_arabic, folder_path):
    transactionFee_url = "https://tahseel.gov.ae/ManagePortal/report-show/97353334-399a-4613-9097-9cf5dc95c690"
    driver.get(transactionFee_url)
    time.sleep(5)
    print_arabic(f"🌐 الانتقال إلى صفحة تقرير رسوم المعاملات لكافة وسائل الدفع | طرق الدفع العالمية")

    from_date_input2 = driver.find_element(By.XPATH, "//*[@id='kt_content']/div/app-show/div/div[1]/div[5]/div[2]/dx-date-box//input[@type='text']")
    to_date_input2 = driver.find_element(By.XPATH, "//*[@id='kt_content']/div/app-show/div/div[1]/div[6]/div[2]/dx-date-box//input[@type='text']")

    from_date_input2.click()
    from_date_input2.clear()
    time.sleep(0.5)
    from_date_input2.send_keys(from_date_arabic)
    from_date_input2.send_keys(Keys.TAB)

    to_date_input2.click()
    to_date_input2.clear()
    time.sleep(0.5)
    to_date_input2.send_keys(to_date_arabic)
    to_date_input2.send_keys(Keys.TAB)

    print_arabic(f"✅ تم ضبط تواريخ البحث لتقرير كافة وسائل الدفع")
    dropdown_xpath = '//*[@id="kt_content"]/div/app-show/div/div[1]/div[3]/div[2]/dx-tag-box/div[1]/div/div[1]'
    ok_button_xpath = '/html/body/div[2]/div/div[2]/div/div[2]/div[1]/div/div'
    select_payment_methods(driver, ["بطاقة ائتمان", "جهاز الدفع البنكي", "جوجل باي","سامسونج باي","آبل باي"],dropdown_xpath,ok_button_xpath)

    submit_button2 = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="kt_content"]/div/app-show/div/div[2]/button')))
    driver.execute_script("arguments[0].click();", submit_button2)
    print_arabic(f"✅ تم الضغط على زر عرض تقرير  كافة وسائل الدفع")

    export_report_to_excel_fixed(folder_path,'BankPayments')

def AmanatUniversalPayments(driver, from_date_arabic, to_date_arabic, folder_path):
    transactionFee_url = "https://tahseel.gov.ae/ManagePortal/report-show/97353334-399a-4613-9097-9cf5dc95c690"
    driver.get(transactionFee_url)
    time.sleep(5)
    print_arabic(f"🌐 الانتقال إلى صفحة تقرير رسوم المعاملات لكافة وسائل الدفع (أمانات)")
    try:
        wait = WebDriverWait(driver, 15)
        radio_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//div[@class='dx-item-content' and contains(text(), 'معاملات امانات')]/parent::div[@role='radio']")))
                
        if radio_button.is_displayed() and radio_button.is_enabled():
            driver.execute_script("arguments[0].scrollIntoView();", radio_button)
            radio_button.click()
            print_arabic(f"✅ تم الضغط على خيار أمانات")
        else:
            print_arabic(f"❌ لم يتم الضغط على خيار أمانات")

    except Exception as e:
        print_arabic(f"حدث الخطأ التالي:{e}")
            
    from_date_input2 = driver.find_element(By.XPATH, "//*[@id='kt_content']/div/app-show/div/div[1]/div[5]/div[2]/dx-date-box//input[@type='text']")
    to_date_input2 = driver.find_element(By.XPATH, "//*[@id='kt_content']/div/app-show/div/div[1]/div[6]/div[2]/dx-date-box//input[@type='text']")

    from_date_input2.click()
    from_date_input2.clear()
    time.sleep(0.5)
    from_date_input2.send_keys(from_date_arabic)
    from_date_input2.send_keys(Keys.TAB)

    to_date_input2.click()
    to_date_input2.clear()
    time.sleep(0.5)
    to_date_input2.send_keys(to_date_arabic)
    to_date_input2.send_keys(Keys.TAB)

    print_arabic(f"✅ تم ضبط تواريخ البحث لتقرير كافة وسائل الدفع أمانات")

    submit_button2 = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="kt_content"]/div/app-show/div/div[2]/button')))
    driver.execute_script("arguments[0].click();", submit_button2)
    print_arabic(f"✅ تم الضغط على زر عرض تقرير كافة وسائل الدفع أمانات")

    export_report_to_excel_fixed(folder_path,'TransactionFeeReportforAllPaymentMethods(Deposit)')

def TotalCreditCardReport(driver, from_date_arabic, to_date_arabic, folder_path):
    transactionFee_url = "https://tahseel.gov.ae/ManagePortal/report-show/fedfceac-2366-407e-881a-29fa1ec5365b"
    driver.get(transactionFee_url)
    time.sleep(5)
    print_arabic(f"🌐 الانتقال لصفحة تقرير اجمالي بطاقات الائتمان")

    dropdown_xpath = '//*[@id="kt_content"]/div/app-show/div/div[1]/div[8]/div[2]/dx-tag-box/div[1]/div/div[1]'
    ok_button_xpath = '/html/body/div[2]/div/div[2]/div/div[2]/div[1]/div/div'

    select_payment_methods(driver, ["معاملات ايراد", "معاملات امانات"],dropdown_xpath,ok_button_xpath)

    from_date_input2 = driver.find_element(By.XPATH,"//*[@id='kt_content']/div/app-show/div/div[1]/div[5]/div[2]/dx-date-box/div/div/div[1]/input[@type='text']")
    to_date_input2 = driver.find_element(By.XPATH, "//*[@id='kt_content']/div/app-show/div/div[1]/div[6]/div[2]/dx-date-box/div/div/div[1]/input[@type='text']")
                                                        
    from_date_input2.click()
    from_date_input2.clear()
    time.sleep(0.5)
    from_date_input2.send_keys(from_date_arabic)
    from_date_input2.send_keys(Keys.TAB)

    to_date_input2.click()
    to_date_input2.clear()
    time.sleep(0.5)
    to_date_input2.send_keys(to_date_arabic)
    to_date_input2.send_keys(Keys.TAB)

    print_arabic(f"✅ تم ضبط تواريخ البحث لتقرير إجمالي بطاقات الائتمان")

    submit_button2 = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="kt_content"]/div/app-show/div/div[2]/button')))
    driver.execute_script("arguments[0].click();", submit_button2)
    print_arabic(f"✅ تم الضغط على زر عرض تقرير إجمالي بطاقات الائتمان")
    export_report_to_excel_fixed(folder_path)

def TotalreportOfSmartReceiptPrintingFees(driver, from_date_arabic, to_date_arabic, folder_path):
    transactionFee_url = "https://tahseel.gov.ae/ManagePortal/report-show/6f3f71e3-246e-48ed-853f-b5b6966a5267"
    driver.get(transactionFee_url)
    time.sleep(5)
    print_arabic(f"🌐 الانتقال لصفحة إجمالي رسوم طباعة الإيصال الذكي")

    select_second_option_from_dropdown("//*[@id='kt_content']/div/app-show/div/div[1]/div[6]/div[2]/dx-select-box//div[contains(@class,'dx-dropdowneditor-button')]", "مدفوعة")

    from_date_input2 = driver.find_element(By.XPATH, "//*[@id='kt_content']/div/app-show/div/div[1]/div[3]/div[2]/dx-date-box//input[@type='text']")
    to_date_input2 = driver.find_element(By.XPATH, "//*[@id='kt_content']/div/app-show/div/div[1]/div[4]/div[2]/dx-date-box//input[@type='text']")

    from_date_input2.click()
    from_date_input2.clear()
    time.sleep(0.5)
    from_date_input2.send_keys(from_date_arabic)
    from_date_input2.send_keys(Keys.TAB)

    to_date_input2.click()
    to_date_input2.clear()
    time.sleep(0.5)
    to_date_input2.send_keys(to_date_arabic)
    to_date_input2.send_keys(Keys.TAB)

    print_arabic(f"✅ تم ضبط تواريخ البحث لتقرير الإيصال الذكي")

    submit_button2 = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="kt_content"]/div/app-show/div/div[2]/button')))
    driver.execute_script("arguments[0].click();", submit_button2)
    print_arabic(f"✅ تم الضغط على زر عرض تقرير الإيصال الذكي")
    export_report_to_excel_fixed(folder_path)

def TotalReportOnSupportServicesTransactionsSection(driver, from_date_arabic, to_date_arabic, folder_path):
    transactionFee_url = "https://tahseel.gov.ae/ManagePortal/report-show/b541adc6-ef50-4019-aac2-32e748add600"
    driver.get(transactionFee_url)
    time.sleep(5)
    print_arabic(f"🌐 الانتقال لصفحة تقرير إجمالي معاملات الخدمات الداعمة")

    try:
        wait = WebDriverWait(driver, 15)
        radio_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//div[@class='dx-item-content' and contains(text(), 'معاملات ايراد')]/parent::div[@role='radio']")))
            
        if radio_button.is_displayed() and radio_button.is_enabled():
            driver.execute_script("arguments[0].scrollIntoView();", radio_button)
            radio_button.click()
            print_arabic(f"✅ تم اختيار معاملات الإيراد بنجاح")
        else:
            print_arabic(f"❌ لم يتم اختيار خيار معاملات الإيراد")

    except Exception as e:
        print_arabic(f"حدث الخطأ التالي:{e}")
        
    select_second_option_from_dropdown("//*[@id='kt_content']/div/app-show/div/div[1]/div[11]/div[2]/dx-select-box//div[contains(@class,'dx-dropdowneditor-button')]", "مدفوعة")

    from_date_input4 = driver.find_element(By.XPATH, "//*[@id='kt_content']/div/app-show/div/div[1]/div[7]/div[2]/dx-date-box//input[@type='text']")                                                     
    to_date_input4 = driver.find_element(By.XPATH, "//*[@id='kt_content']/div/app-show/div/div[1]/div[8]/div[2]/dx-date-box//input[@type='text']")

    from_date_input4.click()
    from_date_input4.clear()
    time.sleep(0.5)
    from_date_input4.send_keys(from_date_arabic)
    from_date_input4.send_keys(Keys.TAB)

    to_date_input4.click()
    to_date_input4.clear()
    time.sleep(0.5)
    to_date_input4.send_keys(to_date_arabic)
    to_date_input4.send_keys(Keys.TAB)

    print_arabic(f"✅ تم ضبط تواريخ البحث لتقرير إجمالي الخدمات الداعمة")
    submit_button3 = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="kt_content"]/div/app-show/div/div[2]/button')))
    driver.execute_script("arguments[0].click();", submit_button3)
    print_arabic(f"✅ تم الضغط على زر عرض تقرير إجمالي الخدمات الداعمة")
    export_report_to_excel_fixed(folder_path)

def TotalChargesReportbyRevenueSourceSection(driver, from_date_arabic, to_date_arabic, folder_path):
    transactionFee_url = "https://tahseel.gov.ae/ManagePortal/report-show/cb6cde66-44d4-4755-84b5-ec32e76c3d30"
    driver.get(transactionFee_url)
    time.sleep(5)
    print_arabic(f"🌐 الانتقال لتقرير إجمالي التحملات حسب جهة الإيراد")
                
    select_second_option_from_dropdown("//*[@id='kt_content']/div/app-show/div/div[1]/div[7]/div[2]/dx-select-box//div[contains(@class,'dx-dropdowneditor-button')]", "مدفوعة")
        
    select_second_option_from_dropdown("//*[@id='kt_content']/div/app-show/div/div[1]/div[6]/div[2]/dx-select-box//div[contains(@class,'dx-dropdowneditor-button')]", "رسوم إيرادات")

    from_date_input4 = driver.find_element(By.XPATH, "//*[@id='kt_content']/div/app-show/div/div[1]/div[3]/div[2]/dx-date-box//input[@type='text']")                                                    
    to_date_input4 = driver.find_element(By.XPATH,   "//*[@id='kt_content']/div/app-show/div/div[1]/div[4]/div[2]/dx-date-box//input[@type='text']")

    from_date_input4.click()
    from_date_input4.clear()
    time.sleep(0.5)
    from_date_input4.send_keys(from_date_arabic)
    from_date_input4.send_keys(Keys.TAB)

    to_date_input4.click()
    to_date_input4.clear()
    time.sleep(0.5)
    to_date_input4.send_keys(to_date_arabic)
    to_date_input4.send_keys(Keys.TAB)

    print_arabic(f"✅ تم ضبط تواريخ البحث في تقرير إجمالي التحملات حسب جهة الإيراد")

    submit_button2 = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="kt_content"]/div/app-show/div/div[2]/button')))
    driver.execute_script("arguments[0].click();", submit_button2)
    print_arabic(f"✅ تم الضغط على زر عرض تقرير إجمالي التحملات حسب جهة الإيراد")
    export_report_to_excel_fixed(folder_path)

def TotalTaxReportSection(driver, from_date_arabic, to_date_arabic, folder_path):
    transactionFee_url = "https://tahseel.gov.ae/ManagePortal/report-show/366f4450-11cf-4b44-a5b2-66c472dbe3c1"
    driver.get(transactionFee_url)
    time.sleep(5)
    print_arabic(f"🌐 الانتقال لتقرير إجمالي بالرسوم الضريبية")

    select_second_option_from_dropdown("//*[@id='kt_content']/div/app-show/div/div[1]/div[9]/div[2]/dx-select-box//div[contains(@class,'dx-dropdowneditor-button')]", "مدفوعة")

    try:
        wait = WebDriverWait(driver, 15)

        radio_button2 = wait.until(EC.element_to_be_clickable((By.XPATH,"//div[@class='dx-item-content' and contains(text(), 'معاملات ايراد')]/parent::div[@role='radio']")))
            
        if radio_button2.is_displayed() and radio_button2.is_enabled():
            driver.execute_script("arguments[0].scrollIntoView();", radio_button2)
            radio_button2.click()
            print_arabic(f"✅ تم اختيار خيار معاملات الإيراد")
        else:
            print_arabic(f"❌ لم يتم اختيار خيار معاملات الإيراد")

    except Exception as e:
        print_arabic(f"حدث الخطأ التالي في تقرير إجمالي الرسوم الضريبية{e}")
        
    try:
        from_date_input5 = driver.find_element(By.XPATH, "//*[@id='kt_content']/div/app-show/div/div[1]/div[5]/div[2]/dx-date-box//input[@type='text']")
        to_date_input5 = driver.find_element(By.XPATH,   "//*[@id='kt_content']/div/app-show/div/div[1]/div[6]/div[2]/dx-date-box//input[@type='text']")
    except Exception as e:
        print_arabic(f"لم يتم العثور على حقل تاريخ من وتاريخ إلى في تقرير إجمالي الرسوم الضريبية{e}")

    from_date_input5.click()
    from_date_input5.clear()
    time.sleep(0.5)
    from_date_input5.send_keys(from_date_arabic)
    from_date_input5.send_keys(Keys.TAB)

    to_date_input5.click()
    to_date_input5.clear()
    time.sleep(0.5)
    to_date_input5.send_keys(to_date_arabic)
    to_date_input5.send_keys(Keys.TAB)
    print_arabic(f"✅ تم ضبط تواريخ البحث في تقرير إجمالي الرسوم الضريبية")
        
    submit_button3 = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="kt_content"]/div/app-show/div/div[2]/button')))
    driver.execute_script("arguments[0].click();", submit_button3)
    print_arabic(f"✅ تم الضغط على زر عرض تقرير إجمالي الرسوم الضريبية")
    export_report_to_excel_fixed(folder_path)

def TotalTransactionReportSection(driver, from_date_arabic, to_date_arabic, folder_path):
    transactionFee_url = "https://tahseel.gov.ae/ManagePortal/report-show/132a8205-691c-4c1d-92f5-5c507020940e"
    driver.get(transactionFee_url)
    time.sleep(5)
    print_arabic(f"🌐 الانتقال إلى تقرير إجمالي بالمعاملات")

    select_second_option_from_dropdown("//*[@id='kt_content']/div/app-show/div/div[1]/div[10]/div[2]/dx-select-box//div[contains(@class,'dx-dropdowneditor-button')]", "مدفوعة")

    try:
        wait = WebDriverWait(driver, 15)
        radio_button2 = wait.until(EC.element_to_be_clickable((By.XPATH,"//div[@class='dx-item-content' and contains(text(), 'معاملات ايراد')]/parent::div[@role='radio']")))
            
        if radio_button2.is_displayed() and radio_button2.is_enabled():
            driver.execute_script("arguments[0].scrollIntoView();", radio_button2)
            radio_button2.click()
            print_arabic(f"✅ تم اختيار خيار معاملات الإيراد")
        else:
            print_arabic(f"❌ لم يتم اختيار خيار معاملات الإيراد")

    except Exception as e:
        print_arabic(f"حدثت المشكلة التالية في تقرير إجمالي بالمعاملات{e}")
        
    try:
        from_date_input5 = driver.find_element(By.XPATH, "//*[@id='kt_content']/div/app-show/div/div[1]/div[5]/div[2]/dx-date-box//input[@type='text']")
        to_date_input5 = driver.find_element(By.XPATH, "//*[@id='kt_content']/div/app-show/div/div[1]/div[6]/div[2]/dx-date-box//input[@type='text']")
    except Exception as e:
        print_arabic(f"لم يتم العثور على حقول التواريخ في تقرير إجمالي بالمعاملات{e}")

    from_date_input5.click()
    from_date_input5.clear()
    time.sleep(0.5)
    from_date_input5.send_keys(from_date_arabic)
    from_date_input5.send_keys(Keys.TAB)

    to_date_input5.click()
    to_date_input5.clear()
    time.sleep(0.5)
    to_date_input5.send_keys(to_date_arabic)
    to_date_input5.send_keys(Keys.TAB)
    print_arabic(f"✅ تم ضبط تواريخ البحث في تقرير إجمالي بالمعاملات")
        
    submit_button5 = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="kt_content"]/div/app-show/div/div[2]/button')))
    driver.execute_script("arguments[0].click();", submit_button5)
    print_arabic(f"✅ تم الضغط على زر عرض التقرير في تقرير إجمالي بالمعاملات")
    export_report_to_excel_fixed(folder_path)

def TransactionPaymentServicesSummaryDepositReceivableReportSection(driver, from_date_arabic, to_date_arabic, folder_path):
    transactionFee_url = "https://tahseel.gov.ae/ManagePortal/report-show/962e3249-71d7-4dc9-973d-da2005ae7745"
    driver.get(transactionFee_url)
    time.sleep(5)
    print_arabic(f"🌐 الانتقال إلى تقـريـر إجمالى بالمعاملات حسب جهة الامانة - المقبوضات")                                                                         
        
        
    try:
        from_date_input = driver.find_element(By.XPATH, "//*[@id='kt_content']/div/app-show/div/div[1]/div[7]/div[2]/dx-date-box//input[@type='text']")
        to_date_input = driver.find_element(By.XPATH, "//*[@id='kt_content']/div/app-show/div/div[1]/div[8]/div[2]/dx-date-box//input[@type='text']")
                                                       
    except Exception as e:
        print_arabic(f"لم يتم العثور على حقول التواريخ في تقرير إجمالى بالمعاملات حسب جهة الامانة - المقبوضات{e}")

    from_date_input.click()
    from_date_input.clear()
    time.sleep(0.5)
    from_date_input.send_keys(from_date_arabic)
    from_date_input.send_keys(Keys.TAB)

    to_date_input.click()
    to_date_input.clear()
    time.sleep(0.5)
    to_date_input.send_keys(to_date_arabic)
    to_date_input.send_keys(Keys.TAB)
    print_arabic(f"✅ تم ضبط تواريخ البحث في تقرير إجمالى بالمعاملات حسب جهة الامانة - المقبوضات")
        
    submit_button5 = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="kt_content"]/div/app-show/div/div[2]/button')))
    driver.execute_script("arguments[0].click();", submit_button5)
    print_arabic(f"✅ تم الضغط على زر عرض التقرير في تقرير إجمالى بالمعاملات حسب جهة الامانة - المقبوضات")
    export_report_to_excel_fixed(folder_path)

def run_total_report_with_retry(expected_name, folder_path, report_func, retries=3, *args, **kwargs):
    
    expected_file = os.path.join(folder_path, expected_name)
    attempt = 0

    while attempt < retries:
        attempt += 1
        print_arabic(f"🔄 محاولة رقم {attempt} لتحميل التقرير: {expected_name}")

        # Call the provided report function
        report_func(*args, **kwargs)

        # Give some time for file to download
        time.sleep(3)

        # Check if file exists
        if os.path.exists(expected_file):
            print_arabic(f"✅ تم العثور على الملف: {expected_name}")
            return expected_file
        else:
            print_arabic(f"❌ الملف {expected_name} غير موجود، إعادة المحاولة ...")

    print_arabic(f"⛔ فشل تحميل الملف {expected_name} بعد {retries} محاولات")
    return None

def get_latest_otp_email(subject_keyword, min_received_time=None, max_days_lookback=None):
    outlook = win32com.client.gencache.EnsureDispatch("Outlook.Application").GetNamespace("MAPI")
    # outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")

    cutoff = None
    if max_days_lookback is not None:
        try:
            cutoff = datetime.now() - timedelta(days=float(max_days_lookback))
        except Exception:
            cutoff = None

    messages_with_time = []

    def collect_from_folder(folder):
        try:
            items = folder.Items
            count = items.Count
            for i in range(1, count + 1):
                try:
                    msg = items.Item(i)
                except Exception:
                    continue

                try:
                    mc = getattr(msg, "MessageClass", "") or ""
                    if not mc.startswith("IPM.Note"):
                        continue
                except Exception:
                    continue

                try:
                    rcv = getattr(msg, "ReceivedTime", None)
                except Exception:
                    rcv = None
                if rcv is None:
                    continue

                # Skip old messages if cutoff provided
                if cutoff is not None:
                    try:
                        if rcv < cutoff:
                            continue
                    except Exception:
                        pass

                # ✅ NEW: skip messages received before OTP was sent
                if min_received_time is not None:
                    try:
                        if rcv < min_received_time:
                            continue
                    except Exception:
                        pass

                messages_with_time.append((rcv, msg))
        except Exception:
            pass

        try:
            sub_count = folder.Folders.Count
            for j in range(1, sub_count + 1):
                try:
                    collect_from_folder(folder.Folders.Item(j))
                except Exception:
                    continue
        except Exception:
            pass

    try:
        stores_count = outlook.Stores.Count
        for s in range(1, stores_count + 1):
            try:
                store = outlook.Stores.Item(s)
                root = store.GetRootFolder()
                collect_from_folder(root)
            except Exception:
                continue
    except Exception:
        try:
            default_root = outlook.Folders.Item(1)
            collect_from_folder(default_root)
        except Exception:
            pass

    if not messages_with_time:
        print("⚠️ No mail items found after OTP request.")
        return None

    messages_with_time.sort(key=lambda t: t[0], reverse=True)

    keyword_lower = subject_keyword.lower()
    for rcv, msg in messages_with_time:
        try:
            subj = getattr(msg, "Subject", "") or ""
            if keyword_lower in subj.lower():
                try:
                    body = getattr(msg, "Body", "") or getattr(msg, "HTMLBody", "") or ""
                except Exception:
                    body = ""
                m = re.search(r"\b(\d{6})\b", body)
                if m:
                    otp = m.group(1)
                    print(f"✅ Found OTP in '{subj}' (received {rcv}): {otp}")
                    return otp
        except Exception:
            continue

    print("⚠️ No matching OTP email found after login time.")
    return None

def load_credentials(filename="creds.json"):
    script_dir = os.path.dirname(os.path.abspath(__file__))
    path = os.path.join(script_dir, filename)
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)

    fernet = Fernet(FERNET_KEY)
    username_input = fernet.decrypt(data["encrypted_username"].encode()).decode()
    password_input = fernet.decrypt(data["encrypted_password"].encode()).decode()

    return username_input, password_input

def parse_flexible_date(date_str):
 
    date_str = date_str.strip().replace("-", "/").replace(".", "/")
    
    # Handle 8-digit number like 18092025
    if date_str.isdigit() and len(date_str) == 8:
        date_str = f"{date_str[:2]}/{date_str[2:4]}/{date_str[4:]}"
    
    try:
        return datetime.strptime(date_str, "%d/%m/%Y")
    except ValueError:
        print_arabic("تنسيق التاريخ غير صحيح")


################################################################################################
while True:
    warnings.filterwarnings("ignore")
    try :

        from_input  = input("📅 Enter FROM date (dd/mm/yyyy): ")
        to_input  = input("📅 Enter TO date (dd/mm/yyyy): ")

        from_date = parse_flexible_date(from_input)
        to_date = parse_flexible_date(to_input)
        today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

        # from_date_str = input("📅 Enter FROM date (dd/mm/yyyy): ").strip() +" 12:00 ص"
        # to_date_str = input("📅 Enter TO date (dd/mm/yyyy): ").strip() +" 11:59 م"
                
        # from_date_str_fixed = from_date_str.replace("ص", "AM").replace("م", "PM")
        # to_date_str_fixed = to_date_str.replace("ص", "AM").replace("م", "PM")
        if from_date > today or to_date > today:
            print_arabic("⚠️ لا يمكن إدخال تواريخ في المستقبل. حاول مرة أخرى.")
            continue

        if from_date > to_date:
            print_arabic("⚠️ تاريخ البداية لا يجب أن يكون بعد تاريخ النهاية. حاول مرة أخرى.")
            continue

        from_date_str = from_date.strftime("%d/%m/%Y") + " 12:00 ص"
        to_date_str = to_date.strftime("%d/%m/%Y") + " 11:59 م"

        from_date_str_fixed = from_date_str.replace("ص", "AM").replace("م", "PM")
        to_date_str_fixed = to_date_str.replace("ص", "AM").replace("م", "PM")
        try:
            from_date_obj = datetime.strptime(from_date_str_fixed, "%d/%m/%Y %I:%M %p")
            to_date_obj = datetime.strptime(to_date_str_fixed, "%d/%m/%Y %I:%M %p")
            # today = datetime.now()

            # If valid, format for Arabic display
            from_date_arabic = from_date_obj.strftime("%d/%m/%Y %I:%M %p").replace("AM", "ص").replace("PM", "م")
            to_date_arabic = to_date_obj.strftime("%d/%m/%Y %I:%M %p").replace("AM", "ص").replace("PM", "م")
            
        except ValueError:
            print_arabic("⚠️ تنسيق التاريخ غير صحيح. الرجاء إدخال التاريخ بصيغة صحيحة مثل 27/05/2025.")

        # Setup folder first
        download_folder = os.path.expanduser("~/Downloads")
        download_folder = os.path.abspath(download_folder)
        folder_name = f"{from_date_obj.strftime('%Y-%m-%d')}-{to_date_obj.strftime('%Y-%m-%d')}"
        folder_path = os.path.join(download_folder, folder_name)
        os.makedirs(folder_path, exist_ok=True)
        # region Setup the App
        # Setup WebDriver
        options = webdriver.ChromeOptions()
        chrome_options = Options()

        #########################################################
        # # Add options to mimic human
        # options.add_argument("--disable-blink-features=AutomationControlled")
        # options.add_argument("--start-maximized")
        # options.add_argument("--disable-infobars")
        # options.add_experimental_option("excludeSwitches", ["enable-automation"])
        # options.add_experimental_option('useAutomationExtension', False)
        # # Set a common user-agent
        # options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36")
        # prefs = {
        # "download.default_directory": folder_path,
        # "download.prompt_for_download": False,
        # "download.directory_upgrade": True,
        # "safebrowsing.enabled": True
            # }
        #########################################################
        chrome_options.add_argument("--log-level=3")
        chrome_options.add_argument("--disable-logging")
        chrome_options.add_experimental_option("excludeSwitches", ["enable-logging"])
        
        options.add_argument("--disable-blink-features=AutomationControlled")
        options.add_argument("--window-size=1920,1080")
        options.add_argument("--disable-infobars")
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option('useAutomationExtension', False)
        options.add_argument("--headless=new")

        options.add_argument(
                "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
            )

        prefs = {
                "download.default_directory": folder_path,
                "download.prompt_for_download": False,
                "download.directory_upgrade": True,
                "safebrowsing.enabled": True
            }
        
        options.add_experimental_option("prefs", prefs)

        
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

        # Remove navigator.webdriver flag using JS
        driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
        "source": """
            Object.defineProperty(navigator, 'webdriver', {
            get: () => undefined
            });
        """
        })

        #Remove all files in the folder
        # clear_folder(folder_path)  
        # ==========================
        # region Login Section
        # ==========================
        try:
            FERNET_KEY = b'vZErE-9uFJn2z_CVw6xPfI_Wp0rNUSIEkyMyAktwLBM='  
            login_successful = False
            driver.get("https://tahseel.gov.ae/ManagePortal/auth/login?returnUrl=%2FManagePortal%2F")
            expected_report_url = "https://tahseel.gov.ae/ManagePortal/dashboard"
            wait = WebDriverWait(driver, 15)
            username_input, password_input = load_credentials()

            username = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="kt_login"]/div/div/div/div[2]/form/div[1]/input')))
            password = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="kt_login"]/div/div/div/div[2]/form/div[2]/input')))

            username.clear()
            username.send_keys(username_input)
            password.clear()
            password.send_keys(password_input)
            password.send_keys(Keys.RETURN)

            sendDateTime = datetime.now()

            try:
                error_element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//div[contains(text(),'البيانات المدخلة غير صحيحة')]")))   
                print_arabic(f"❌ فشل تسجيل الدخول: البيانات المدخلة غير صحيحة.")

            except TimeoutException:
                try:
                    # Try to fetch OTP automatically from Outlook
                    otp = get_latest_otp_email("كلمة المرور لمرة واحدة", min_received_time=sendDateTime)
                    if otp:
                        print_arabic(f"🔐 تم استخراج رمز التحقق تلقائياً: {otp}")
                    else:
                        print_arabic(f"⚠️ لم يتم العثور على رمز التحقق في البريد الإلكتروني، الرجاء إدخاله يدوياً.")
                        otp = input("🔢 أدخل رمز التحقق يدوياً: ")

                    otp_verified = False
                    for attempt in range(3):
                        # try to refresh OTP automatically on subsequent attempts if not provided
                        if attempt > 0 and not otp:
                            otp = get_latest_otp_email("كلمة المرور لمرة واحدة", min_received_time=sendDateTime)
                            if otp:
                                print_arabic(f"🔐 تم العثور على رمز تحقق جديد: {otp}")
                            else:
                                otp = input(f"🔢 محاولة {attempt + 1} - أدخل رمز التحقق يدوياً: ")

                        # sanity check
                        if not otp or len(otp) < 6:
                            print_arabic("⚠️ رمز التحقق غير صالح أو قصير. الرجاء المحاولة يدوياً.")
                            otp = input("🔢 أدخل رمز التحقق يدوياً: ")
                            if not otp or len(otp) < 6:
                                # allow loop to continue and possibly re-fetch
                                otp = None
                                continue

                        # Ensure first field exists and is focused
                        first_field_id = "otp-digit-0"
                        try:
                            first_elem = wait.until(EC.presence_of_element_located((By.ID, first_field_id)))
                            try:
                                # scroll, click, and focus the first field
                                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", first_elem)
                                first_elem.click()
                                driver.execute_script("arguments[0].focus();", first_elem)
                            except (ElementClickInterceptedException, WebDriverException):
                                # fallback to JS focus if click fails
                                driver.execute_script("arguments[0].focus();", first_elem)
                            time.sleep(0.12)
                        except Exception:
                            print_arabic("⚠️ الحقول غير موجودة أو غير قابلة للتفاعل الآن.")
                            otp = None
                            continue

                        # Clear all fields via JS and dispatch input event so app reacts
                        for i in range(6):
                            field_id = f"otp-digit-{i}"
                            try:
                                elem = driver.find_element(By.ID, field_id)
                                driver.execute_script(
                                    "arguments[0].value = ''; arguments[0].dispatchEvent(new Event('input', {bubbles:true}));",
                                    elem
                                )
                            except Exception:
                                # ignore if a particular field is absent; we'll try to type into what's present
                                pass
                        time.sleep(0.12)

                        # Type digits robustly, click/focus each field, send_keys, and dispatch input event
                        typed_ok = True
                        for i, ch in enumerate(otp[:6]):
                            field_id = f"otp-digit-{i}"
                            try:
                                elem = wait.until(EC.element_to_be_clickable((By.ID, field_id)))
                                try:
                                    # click then focus - improves chances caret is set
                                    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", elem)
                                    elem.click()
                                except (ElementClickInterceptedException, WebDriverException, StaleElementReferenceException):
                                    # fallback to focusing via JS
                                    driver.execute_script("arguments[0].focus();", elem)
                                    time.sleep(0.05)

                                # clear again via element API to ensure it's empty
                                try:
                                    elem.clear()
                                except Exception:
                                    pass

                                # Primary: try send_keys
                                try:
                                    elem.send_keys(ch)
                                    # Trigger input event in case the site listens for it
                                    driver.execute_script("arguments[0].dispatchEvent(new Event('input', {bubbles:true}));", elem)
                                except Exception:
                                    # Fallback: set value by JS and dispatch events
                                    driver.execute_script(
                                        "arguments[0].value = arguments[1]; arguments[0].dispatchEvent(new Event('input', {bubbles:true}));",
                                        elem, ch
                                    )

                                time.sleep(0.15)
                            except Exception as e:
                                # if a digit fails to type, mark and break so we can retry next attempt
                                print(f"⚠️ خطأ أثناء إدخال الحقل {field_id}: {e}")
                                typed_ok = False
                                break

                        # Small pause for the page to process OTP input (verification)
                        time.sleep(1.2)

                        # Optional: trigger blur or submit if the UI requires it (uncomment if needed)
                        # try:
                        #     driver.execute_script("document.getElementById('otp-digit-5').blur();")
                        # except Exception:
                        #     pass

                        # Wait briefly then check success
                        time.sleep(3)
                        if expected_report_url in driver.current_url:
                            print_arabic("✅ رمز التحقق صحيح، تم تسجيل الدخول بنجاح.")
                            otp_verified = True
                            break
                        else:
                            print_arabic("❌ رمز التحقق غير صحيح. الرجاء المحاولة مرة أخرى.")
                            # reset otp so we try to fetch again next loop
                            otp = None

                    if not otp_verified:
                        print_arabic("🚫 الفشل في التحقق من رمز التحقق بعد 3 محاولات.")
                        driver.quit()
                        exit()
                    else:
                        login_successful = True

                except TimeoutException:
                    print_arabic("⚠️ لم يظهر لا خطأ ولا شاشة التحقق. تحقق من الاتصال أو تغييرات في الموقع.")
        except Exception as e:
            print(f"Error during input: {e}")
        
        if login_successful:
            ################################################################################################################################

             ################################################################################
            # تقرير رسوم المعاملات لكافة وسائل الدفع (أمانات)
            # run_total_report_with_retry("TransactionFeeReportforAllPaymentMethods(Deposit).xlsx",folder_path,AmanatUniversalPayments,3,driver, from_date_arabic, to_date_arabic, folder_path)

            #تقـريـر إجمالى بالمعاملات حسب جهة الامانة - المقبوضات
            # Total Report of Transactions by Revenue Source - Receivables Section
            # ==========================

            # run_total_report_with_retry("transactionpaymentservicessummaryDepositreceivable.xlsx",folder_path,TransactionPaymentServicesSummaryDepositReceivableReportSection,3,driver, from_date_arabic, to_date_arabic, folder_path)
            ################################################################################################################################
            #تقـريـر  إجمالى بالمعاملات حسب جهة الايراد - المقبوضات
            #Total Report of Transactions by Revenue Source - Receivables Section
            # ==========================

            # run_total_report_with_retry("transactionpaymentservicessummaryreceivable_sec.xlsx",folder_path,TotalReportOfTransactionsbyRevenueSourceRec,3,driver, from_date_arabic, to_date_arabic, folder_path)
            ################################################################################
            # تقرير رسوم المعاملات لكافة وسائل الدفع
            # Transactions Fee Report Section
            # ==========================
            # run_total_report_with_retry("TransactionFeesForAllPaymentMethods.xlsx",folder_path,TransactionsFeeReport(driver,from_date_arabic,to_date_arabic,folder_path))
            # run_total_report_with_retry("TransactionFeesForAllPaymentMethods.xlsx",folder_path,TransactionsFeeReport,3,driver, from_date_arabic, to_date_arabic, folder_path)

            ################################################################################
            # تقرير رسوم المعاملات لكافة وسائل الدفع (الدفع بالطرق العالمية)
            # Universal Payments
            # run_total_report_with_retry("BankPayments.xlsx",folder_path,UniversalPayments,3,driver, from_date_arabic, to_date_arabic, folder_path)

            ################################################################################
            # تقرير اجمالي بطاقات الائتمان
            # ShjCreditCardSummery.xlsx
            # ==========================
            # run_total_report_with_retry("ShjCreditCardSummery.xlsx",folder_path,TotalCreditCardReport,3,driver, from_date_arabic, to_date_arabic, folder_path)
            
            ################################################################################
            # تقرير إجمالي لرسوم طباعة الإيصال الذكي
            # Total report of Smart Receipt Details
            # ==========================
            # run_total_report_with_retry("GITFees_ShjGovTransStatistics.xlsx",folder_path,TotalreportOfSmartReceiptPrintingFees,3,driver, from_date_arabic, to_date_arabic, folder_path)

            ################################################################################
            # ==========================
            #تقـريـر  إجمالي بمعاملات الخدمات الداعمة
            #Total report on Support Services Reports Section
            # =========================
            # run_total_report_with_retry("TRANSACTIONPAYMENTDEPENDANTSERVICESSUMMARY_sec.xlsx",folder_path,TotalReportOnSupportServicesTransactionsSection,3,driver, from_date_arabic, to_date_arabic, folder_path)

            ################################################################################
            # ==========================
            #تقرير إجمالي التحملات حسب جهة الإيراد
            #Total Charges Report by Revenue Source Section
            # ==========================
            # run_total_report_with_retry("SummaryReport_of_IncurredFees PerRevenueEntity.xlsx",folder_path,TotalChargesReportbyRevenueSourceSection,3,driver, from_date_arabic, to_date_arabic, folder_path)

            ################################################################################
            # ==========================
            #تقـريـر إجمالي بالرسوم الضريبية
            #Total tax report Section
            # ==========================
            # run_total_report_with_retry("TRANSACTIONTAXSUMARY.xlsx",folder_path,TotalTaxReportSection,3,driver, from_date_arabic, to_date_arabic, folder_path)
            
            ################################################################################
            # ==========================
            # تقرير إجمالي بالمعاملات
            # Total Transaction Report Section
            # ==========================
            # run_total_report_with_retry("ShjGovTransSummary_sec.xlsx",folder_path,TotalTransactionReportSection,3,driver, from_date_arabic, to_date_arabic, folder_path)
            
        ##############################################################################################################################
        # region Getting Values From Excel Files Section
        ###############################################################################################################################
            # تقـريـر  إجمالى بالمعاملات حسب جهة الامانة - المقبوضات
            # transactionpaymentservicessummaryDepositreceivable
            FileName = "transactionpaymentservicessummaryDepositreceivable"
            latest_file_transactionpaymentservicessummaryDepositreceivableFileName = get_latest_excel_file(folder_path, FileName)
            sheet = 'transactionpaymentservicessumma'

            last_FeeValue_Depositreceivable = get_last_value_in_column(latest_file_transactionpaymentservicessummaryDepositreceivableFileName, sheet, ['T','U','V','W','X','Y','Z'])
            print_arabic(f"إجمالي قيمة الرسوم في تقرير إجمالي بالمعاملات حسب جهة الأمانة - المقبوضات:  {last_FeeValue_Depositreceivable}")
            SaveToExcel(f"إجمالي قيمة الرسوم في تقرير إجمالي بالمعاملات حسب جهة الأمانة - المقبوضات" , latest_file_transactionpaymentservicessummaryDepositreceivableFileName,folder_path)
            #############################################################
            #تقـريـر  إجمالى بالمعاملات حسب جهة الايراد - المقبوضات    
            #transactionpaymentservicessummaryreceivable_sec   
            FileName = "transactionpaymentservicessummaryreceivable_sec"
            latest_file_TransactionsbyRevenueSourceReportFileName = get_latest_excel_file(folder_path, FileName)
            sheet = 'transactionpaymentservicessumma'

            last_transaction_value_Revenue = get_last_value_in_column(latest_file_TransactionsbyRevenueSourceReportFileName, sheet,['S','T','U','V','W','X','Y'])
            print_arabic(f"إجمالي رسوم المعاملة في تقرير المقبوضات:  {last_transaction_value_Revenue}")
            SaveToExcel(f"إجمالي رسوم المعاملة في تقرير المقبوضات" , last_transaction_value_Revenue,folder_path)
            
            last_Tax_Total_Revenue = get_last_value_in_column(latest_file_TransactionsbyRevenueSourceReportFileName, sheet,'R')
            print_arabic(f"إجمالي الضريبة في تقرير المقبوضات:  {last_Tax_Total_Revenue}")
            SaveToExcel(f"إجمالي الضريبة في تقرير المقبوضات",last_Tax_Total_Revenue,folder_path)


            # #############################################################
            #كافة وسائل الدفع   
            # TransactionFeesForAllPaymentMethods
            FileName = "TransactionFeesForAllPaymentMethods"
            sheet = 'TransactionFeesForAllPaymentMet'
            latest_file_TransactionFeesForAllPaymentMethodsFileName = get_latest_excel_file(folder_path, FileName)

            last_transaction_value = get_last_value_in_column(latest_file_TransactionFeesForAllPaymentMethodsFileName, sheet,'AO')
            print_arabic(f"إجمالي رسوم المعاملة في تقرير كافة وسائل الدفع:  {last_transaction_value}")
            SaveToExcel(f"إجمالي رسوم المعاملة في تقرير كافة وسائل الدفع",last_transaction_value,folder_path)

            last_tax_value = get_last_value_in_column(latest_file_TransactionFeesForAllPaymentMethodsFileName, sheet,'AN')
            print_arabic(f"إجمالي الضريبة في تقرير كافة وسائل الدفع:  {last_tax_value}")
            SaveToExcel(f"إجمالي الضريبة في تقرير كافة وسائل الدفع",last_tax_value,folder_path)


            last_Tahseel_Fee = get_last_value_in_column(latest_file_TransactionFeesForAllPaymentMethodsFileName, sheet,['AF','AG','AH','AI'])
            print_arabic(f"إجمالي رسوم الخدمات - تحصيل في تقرير كافة وسائل الدفع :  {last_Tahseel_Fee}")
            SaveToExcel(f"إجمالي رسوم الخدمات - تحصيل في تقرير كافة وسائل الدفع",last_Tahseel_Fee,folder_path)

            last_SFD_ServiceFee = get_last_value_in_column(latest_file_TransactionFeesForAllPaymentMethodsFileName, sheet,['AJ','AK','AL'])
            print_arabic(f"إجمالي رسوم الخدمات - المالية المركزية في تقرير كافة وسائل الدفع: {last_SFD_ServiceFee}")
            SaveToExcel(f"إجمالي رسوم الخدمات - المالية المركزية في تقرير كافة وسائل الدفع",last_SFD_ServiceFee,folder_path)

            last_ServiceFee_Charges = get_last_value_in_column(latest_file_TransactionFeesForAllPaymentMethodsFileName, sheet,"W")
            print_arabic(f"إجمالي رسوم الخدمات - تحمل الرسوم في تقرير كافة وسائل الدفع : {last_ServiceFee_Charges}")
            SaveToExcel(f"إجمالي رسوم الخدمات - تحمل الرسوم في تقرير كافة وسائل الدفع",last_ServiceFee_Charges,folder_path)

            last_BankFees = get_last_value_in_column(latest_file_TransactionFeesForAllPaymentMethodsFileName, sheet,["M","N","O"])
            print_arabic(f"إجمالي الرسوم البنكية - تحمل الرسوم البنكية  في تقرير كافة وسائل الدفع: {last_BankFees}")
            SaveToExcel(f"إجمالي الرسوم البنكية - تحمل الرسوم البنكية  في تقرير كافة وسائل الدفع",last_BankFees,folder_path)

            last_VAT_Charge = get_last_value_in_column(latest_file_TransactionFeesForAllPaymentMethodsFileName, sheet,"V")
            print_arabic(f"إجمالي تحمل الضريبة في تقرير كافة وسائل الدفع: {last_VAT_Charge}")
            SaveToExcel(f"إجمالي تحمل الضريبة في تقرير كافة وسائل الدفع",last_VAT_Charge,folder_path)

            last_BankFees_VAT = get_last_value_in_column(latest_file_TransactionFeesForAllPaymentMethodsFileName, sheet,"L")
            print_arabic(f":إجمالي الرسوم البنكية - ضريبة تحمل الرسوم البنكية في تقرير كافة وسائل الدفع {last_BankFees_VAT}")
            SaveToExcel(f"إجمالي الرسوم البنكية - ضريبة تحمل الرسوم البنكية في تقرير كافة وسائل الدفع",last_BankFees_VAT,folder_path)

            Service_Vat_Tahseel_Share = get_last_value_in_column(latest_file_TransactionFeesForAllPaymentMethodsFileName, sheet,"AE")
            print_arabic(f"إجمالي ضريبة تحصيل في تقرير رسوم المعاملات لكافة وسائل الدفع : {Service_Vat_Tahseel_Share}")
            SaveToExcel(f"إجمالي ضريبة تحصيل في تقرير رسوم المعاملات لكافة وسائل الدفع",Service_Vat_Tahseel_Share,folder_path)

            Resacrch_Fees = get_last_value_in_column(latest_file_TransactionFeesForAllPaymentMethodsFileName, sheet,"U")
            print_arabic(f"إجمالي رسوم البحث العلمي من تقرير كافة وسائل الدفع: {Resacrch_Fees}")
            SaveToExcel(f"إجمالي رسوم البحث العلمي من تقرير كافة وسائل الدفع",Resacrch_Fees,folder_path)

            #############################################################
            ########Universal Payment methods########
            FileName = "BankPayments"
            sheet = 'TransactionFeesForAllPaymentMet'
            latest_file_Uni_TransactionFeesForAllPaymentMethodsFileName = get_latest_excel_file(folder_path, FileName)

            bankPayments_NetFees = get_last_value_in_column(latest_file_Uni_TransactionFeesForAllPaymentMethodsFileName, sheet,["A","B","C","D","E"])
            print_arabic(f"إجمالي صافي الرسوم في تقرير كافة وسائل الدفع (المحافظ العالمية) : {bankPayments_NetFees}")
            SaveToExcel(f"إجمالي صافي الرسوم في تقرير كافة وسائل الدفع (المحافظ العالمية)",bankPayments_NetFees,folder_path)

            bankPayments_IncomeFees = get_last_value_in_column(latest_file_Uni_TransactionFeesForAllPaymentMethodsFileName, sheet,"F")
            print_arabic(f"إجمالي ضريبة المدخلات في تقرير كافة وسائل الدفع (المحافظ العالمية) : {bankPayments_IncomeFees}")
            SaveToExcel(f"إجمالي صافي الرسوم في تقرير كافة وسائل الدفع (المحافظ العالمية)",bankPayments_IncomeFees,folder_path)

            bankPayments_BankShare = get_last_value_in_column(latest_file_Uni_TransactionFeesForAllPaymentMethodsFileName, sheet,"G")
            print_arabic(f"إجمالي حصة البنك في تقرير كافة وسائل الدفع (المحافظ العالمية) : {bankPayments_BankShare}")
            SaveToExcel(f"إجمالي حصة البنك في تقرير كافة وسائل الدفع (المحافظ العالمية)",bankPayments_BankShare,folder_path)

            bankPayments_TotalBankFees = get_last_value_in_column(latest_file_Uni_TransactionFeesForAllPaymentMethodsFileName, sheet,["I","J","K"])
            print_arabic(f"إجمالي الرسوم البنكية في تقرير كافة وسائل الدفع (المحافظ العالمية) : {bankPayments_TotalBankFees}")
            SaveToExcel(f"إجمالي الرسوم البنكية في تقرير كافة وسائل الدفع (المحافظ العالمية)",bankPayments_TotalBankFees,folder_path)

            Uni_ServiceFee_SFD = get_last_value_in_column(latest_file_Uni_TransactionFeesForAllPaymentMethodsFileName, sheet,["AJ","AK","AL"])
            SaveToExcel(f"إجمالي رسوم خدمات المالية في تقرير كافة وسائل الدفع (المحافظ العالمية):",Uni_ServiceFee_SFD,folder_path)

            Uni_ServiceFee_Tahseel = get_last_value_in_column(latest_file_Uni_TransactionFeesForAllPaymentMethodsFileName, sheet,["AF","AG","AH","AI"])
            SaveToExcel(f"إجمالي رسوم خدمات تحصيل في تقرير كافة وسائل الدفع (المحافظ العالمية):",Uni_ServiceFee_Tahseel,folder_path)

            Uni_TotalServiceFee = Uni_ServiceFee_SFD + Uni_ServiceFee_Tahseel
            print_arabic(f"إجمالي رسوم خدمات المالية + رسوم خدمات تحصيل في تقرير كافة وسائل الدفع (المحافظ العالمية): {Uni_TotalServiceFee}")
            SaveToExcel(f"إجمالي رسوم خدمات المالية + رسوم خدمات تحصيل في تقرير كافة وسائل الدفع (المحافظ العالمية):",Uni_TotalServiceFee,folder_path)

            Uni_TahseelVAT = get_last_value_in_column(latest_file_Uni_TransactionFeesForAllPaymentMethodsFileName, sheet,"AE")
            print_arabic(f"إجمالي ضريبة تحصيل في تقرير كافة وسائل الدفع (المحافظ العالمية): {Uni_TahseelVAT}")
            SaveToExcel(f"إجمالي ضريبة تحصيل في تقرير كافة وسائل الدفع (المحافظ العالمية):",Uni_TahseelVAT,folder_path)

            Uni_BankFee = get_last_value_in_column(latest_file_Uni_TransactionFeesForAllPaymentMethodsFileName, sheet,["Q","R","S"])
            print_arabic(f"إجمالي المصاريف البنكية في تقرير كافة وسائل الدفع ( وسائل الدفع: المحافظ العالمية): {Uni_BankFee}")
            SaveToExcel(f"إجمالي المصاريف البنكية في تقرير كافة وسائل الدفع ( وسائل الدفع: المحافظ العالمية):",Uni_BankFee,folder_path)

            Uni_BankFee_BearExp = get_last_value_in_column(latest_file_Uni_TransactionFeesForAllPaymentMethodsFileName, sheet,["M","N","O"])
            print_arabic(f"إجمالي تحمل الرسوم البنكية في تقرير وسائل الدفع (المحافظ العالمية): {Uni_BankFee_BearExp}")
            SaveToExcel(f"إجمالي تحمل الرسوم البنكية في تقرير وسائل الدفع (المحافظ العالمية):",Uni_BankFee_BearExp,folder_path)

            Uni_BankFeeAndBearExp = Uni_BankFee + Uni_BankFee_BearExp

            Uni_BankTaxFee = get_last_value_in_column(latest_file_Uni_TransactionFeesForAllPaymentMethodsFileName, sheet,"P")
            print_arabic(f"إجمالي ضريبة المصاريف البنكية في تقرير كافة وسائل الدفع (المحافظ العالمية): {Uni_BankTaxFee}")
            SaveToExcel(f"إجمالي ضريبة المصاريف البنكية في تقرير كافة وسائل الدفع (المحافظ العالمية):",Uni_BankTaxFee,folder_path)

            Uni_BankFeeTotal = get_last_value_in_column(latest_file_Uni_TransactionFeesForAllPaymentMethodsFileName, sheet,"L")
            print_arabic(f"إجمالي ضريبة تحمل الرسوم البنكية في تقرير وسائل الدفع (المحافظ العالمية): {Uni_BankFeeTotal}")
            SaveToExcel(f"إجمالي ضريبة تحمل الرسوم البنكية في تقرير وسائل الدفع (المحافظ العالمية):",Uni_BankFeeTotal,folder_path)
            
            Universal_BankVatFee = Uni_BankTaxFee + Uni_BankFeeTotal

            Uni_BankFeePlusBankFeeVAT = Uni_BankFeeTotal + Uni_BankFee_BearExp		


            Uni_EntityShareFeesForResarchFees = get_last_value_in_column(latest_file_Uni_TransactionFeesForAllPaymentMethodsFileName, sheet,"T")
            print_arabic(f"إجمالي تحمل الجهة للرسوم في تقرير وسائل الدفع (المحافظ العالمية): {Uni_EntityShareFeesForResarchFees}")
            SaveToExcel(f"إجمالي تحمل الجهة للرسوم في تقرير وسائل الدفع (المحافظ العالمية):",Uni_EntityShareFeesForResarchFees,folder_path)

            Uni_ResarchFees = get_last_value_in_column(latest_file_Uni_TransactionFeesForAllPaymentMethodsFileName, sheet,"U")
            print_arabic(f"إجمالي رسوم البحث العلمي في تقرير وسائل الدفع (المحافظ العالمية): {Uni_ResarchFees}")
            SaveToExcel(f"إجمالي رسوم البحث العلمي في تقرير وسائل الدفع (المحافظ العالمية):",Uni_ResarchFees,folder_path)

            Uni_TotalVATForEntityFee = get_last_value_in_column(latest_file_Uni_TransactionFeesForAllPaymentMethodsFileName, sheet,"AN")
            print_arabic(f"إجمالي الضريبة لإجمالي رسوم الجهة في تقرير وسائل الدفع (المحافظ العالمية): {Uni_TotalVATForEntityFee}")
            SaveToExcel(f"إجمالي الضريبة لإجمالي رسوم الجهة في تقرير وسائل الدفع (المحافظ العالمية):",Uni_TotalVATForEntityFee,folder_path)

            Uni_TotalTransactionFeeForEntityFee = get_last_value_in_column(latest_file_Uni_TransactionFeesForAllPaymentMethodsFileName, sheet,"AO")
            print_arabic(f"إجمالي رسوم المعاملة لإجمالي رسوم الجهة في تقرير وسائل الدفع (المحافظ العالمية): {Uni_TotalTransactionFeeForEntityFee}")
            SaveToExcel(f"إجمالي رسوم المعاملة لإجمالي رسوم الجهة في تقرير وسائل الدفع (المحافظ العالمية):",Uni_TotalTransactionFeeForEntityFee,folder_path)
            #############################################################
            # Total Universal Transaction Fee 
            # ShjCreditCardSummery
            FileName = "ShjCreditCardSummery"
            sheet = 'Sheet1'
            latest_file_Uni_ShjCreditCardSummeryFileName = get_latest_excel_file(folder_path, FileName)

            # Uni_TotalTransactionFee = get_last_value_in_column(latest_file_Uni_ShjCreditCardSummeryFileName, sheet,["AK","AL","AM"])
            Uni_TotalTransactionFee = get_last_value_in_column(latest_file_Uni_ShjCreditCardSummeryFileName, sheet,"Y")
            print_arabic(f"إجمالي رسوم المعاملة في تقرير معاملات بطاقات الائتمان: {Uni_TotalTransactionFee}")
            SaveToExcel(f"إجمالي رسوم المعاملة في تقرير معاملات بطاقات الائتمان:",Uni_TotalTransactionFee,folder_path)

            # Uni_TotalTransactionFeeVat = get_last_value_in_column(latest_file_Uni_ShjCreditCardSummeryFileName, sheet,["AI","AJ"])
            Uni_TotalTransactionFeeVat = get_last_value_in_column(latest_file_Uni_ShjCreditCardSummeryFileName, sheet,"X")
            print_arabic(f"إجمالي ضريبة المخرجات لرسوم المعاملة في تقرير معاملات بطاقات الائتمان: {Uni_TotalTransactionFeeVat}")
            SaveToExcel(f"إجمالي ضريبة المخرجات لرسوم المعاملة في تقرير معاملات بطاقات الائتمان:",Uni_TotalTransactionFeeVat,folder_path)

            # Uni_ResearchSupportFees = get_last_value_in_column(latest_file_Uni_ShjCreditCardSummeryFileName, sheet,["AG","AH"])
            Uni_ResearchSupportFees = get_last_value_in_column(latest_file_Uni_ShjCreditCardSummeryFileName, sheet,"W")
            print_arabic(f"إجمالي رسوم دعم الأبحاث في تقرير معاملات بطاقات الائتمان: {Uni_ResearchSupportFees}")
            SaveToExcel(f"إجمالي رسوم دعم الأبحاث في تقرير معاملات بطاقات الائتمان:",Uni_ResearchSupportFees,folder_path)

            # Uni_EntityHold = get_last_value_in_column(latest_file_Uni_ShjCreditCardSummeryFileName, sheet,["AE","AF"])
            Uni_EntityHold = get_last_value_in_column(latest_file_Uni_ShjCreditCardSummeryFileName, sheet,["U","V"])
            print_arabic(f"إجمالي تحمل الجهة لرسوم دعم الأبحاث في تقرير معاملات بطاقات الائتمان: {Uni_EntityHold}")
            SaveToExcel(f"إجمالي تحمل الجهة لرسوم دعم الأبحاث في تقرير معاملات بطاقات الائتمان:",Uni_EntityHold,folder_path)

            # Uni_CreditRepprt_TotalServiceFee = get_last_value_in_column(latest_file_Uni_ShjCreditCardSummeryFileName, sheet,["AA","AB","AC","AD"])
            Uni_CreditRepprt_TotalServiceFee = get_last_value_in_column(latest_file_Uni_ShjCreditCardSummeryFileName, sheet,["S","T"])
            print_arabic(f"إجمالي رسوم الخدمات في تقرير معاملات بطاقات الائتمان: {Uni_CreditRepprt_TotalServiceFee}")
            SaveToExcel(f"إجمالي رسوم الخدمات في تقرير معاملات بطاقات الائتمان:",Uni_CreditRepprt_TotalServiceFee,folder_path)

            # Uni_CreditRepprt_OutputFee = get_last_value_in_column(latest_file_Uni_ShjCreditCardSummeryFileName, sheet,["Y","Z"])
            Uni_CreditRepprt_OutputFee = get_last_value_in_column(latest_file_Uni_ShjCreditCardSummeryFileName, sheet,["P","Q","R"])
            print_arabic(f"إجمالي ضريبة المخرجات لرسوم الخدمات في تقرير معاملات بطاقات الائتمان: {Uni_CreditRepprt_OutputFee}")
            SaveToExcel(f"إجمالي ضريبة المخرجات لرسوم الخدمات في تقرير معاملات بطاقات الائتمان:",Uni_CreditRepprt_OutputFee,folder_path)

            # Uni_EntityHoldServiceFee = get_last_value_in_column(latest_file_Uni_ShjCreditCardSummeryFileName, sheet,["U","V","W","X"])
            Uni_EntityHoldServiceFee = get_last_value_in_column(latest_file_Uni_ShjCreditCardSummeryFileName, sheet,"O")
            print_arabic(f"إجمالي تحمل الجهة لرسوم الخدمات في تقرير معاملات بطاقات الائتمان: {Uni_EntityHoldServiceFee}")
            SaveToExcel(f"إجمالي تحمل الجهة لرسوم الخدمات في تقرير معاملات بطاقات الائتمان:",Uni_EntityHoldServiceFee,folder_path)

            # Uni_CreditRepprt_BankFee = get_last_value_in_column(latest_file_Uni_ShjCreditCardSummeryFileName, sheet,["S","T"])
            Uni_CreditRepprt_BankFee = get_last_value_in_column(latest_file_Uni_ShjCreditCardSummeryFileName, sheet,"N")
            print_arabic(f"إجمالي رسوم البنك في تقرير معاملات بطاقات الائتمان: {Uni_CreditRepprt_BankFee}")
            SaveToExcel(f"إجمالي رسوم البنك في تقرير معاملات بطاقات الائتمان:",Uni_CreditRepprt_BankFee,folder_path)

            # Uni_CreditRepprt_OutputTax = get_last_value_in_column(latest_file_Uni_ShjCreditCardSummeryFileName, sheet,["Q","R"])
            Uni_CreditRepprt_OutputTax = get_last_value_in_column(latest_file_Uni_ShjCreditCardSummeryFileName, sheet,"M")
            print_arabic(f"إجمالي ضريبة المخرجات لرسوم البنك في تقرير معاملات بطاقات الائتمان: {Uni_CreditRepprt_OutputTax}")
            SaveToExcel(f"إجمالي ضريبة المخرجات لرسوم البنك في تقرير معاملات بطاقات الائتمان:",Uni_CreditRepprt_OutputTax,folder_path)

            Uni_BankFeesPlusOutputTax = Uni_CreditRepprt_OutputTax+Uni_CreditRepprt_BankFee

            # Uni_DistributeBankFees_Bank = get_last_value_in_column(latest_file_Uni_ShjCreditCardSummeryFileName, sheet,["K","L"])
            Uni_DistributeBankFees_Bank = get_last_value_in_column(latest_file_Uni_ShjCreditCardSummeryFileName, sheet,"I")
            print_arabic(f"إجمالي حصة البنك في توزيع رسوم البنك في تقرير معاملات بطاقات الائتمان: {Uni_DistributeBankFees_Bank}")
            SaveToExcel(f"إجمالي حصة البنك في توزيع رسوم البنك في تقرير معاملات بطاقات الائتمان:",Uni_DistributeBankFees_Bank,folder_path)

            # Uni_DistributeBankFees_IncomeFee = get_last_value_in_column(latest_file_Uni_ShjCreditCardSummeryFileName, sheet,["G","H","I","J"])
            Uni_DistributeBankFees_IncomeFee = get_last_value_in_column(latest_file_Uni_ShjCreditCardSummeryFileName, sheet,["F","G","H"])
            print_arabic(f"إجمالي ضريبة المدخلات في توزيع رسوم البنك في تقرير معاملات بطاقات الائتمان: {Uni_DistributeBankFees_IncomeFee}")
            SaveToExcel(f"إجمالي ضريبة المدخلات في توزيع رسوم البنك في تقرير معاملات بطاقات الائتمان:",Uni_DistributeBankFees_IncomeFee,folder_path)

            # Uni_DistributeBankFees_EntityHold = get_last_value_in_column(latest_file_Uni_ShjCreditCardSummeryFileName, sheet,["C","D"])
            Uni_DistributeBankFees_EntityHold = get_last_value_in_column(latest_file_Uni_ShjCreditCardSummeryFileName, sheet,"C")
            print_arabic(f"إجمالي تحمل الجهة في توزيع رسوم البنك في تقرير معاملات بطاقات الائتمان: {Uni_DistributeBankFees_EntityHold}")
            SaveToExcel(f"إجمالي تحمل الجهة في توزيع رسوم البنك في تقرير معاملات بطاقات الائتمان:",Uni_DistributeBankFees_EntityHold,folder_path)

            Uni_SFDTotal = get_last_value_in_column(latest_file_Uni_ShjCreditCardSummeryFileName, sheet,["A","B"])
            print_arabic(f"إجمالي المالية في تقرير معاملات بطاقات الائتمان: {Uni_SFDTotal}")
            SaveToExcel(f"إجمالي المالية في تقرير معاملات بطاقات الائتمان:",Uni_SFDTotal,folder_path)

            #############################################################
            
            FileName = "ShjGovTransSummary_sec"
            latest_file_ShjGovTransSummary_secFileName = get_latest_excel_file(folder_path, FileName)
            
            sheet = 'جمارك الشارقة'
            Total_Transactions_all_dep = get_last_value_in_column(latest_file_ShjGovTransSummary_secFileName, sheet,"A")
            print_arabic(f"إجمالي رسوم المعاملات لجميع الدوائر في تقرير أجمالي المعاملات : {Total_Transactions_all_dep}")
            SaveToExcel(f"إجمالي رسوم المعاملات لجميع الدوائر في تقرير أجمالي المعاملات",Total_Transactions_all_dep,folder_path)

            #############################################################
            # TRANSACTIONPAYMENTDEPENDANTSERVICESSUMMARY_sec
            # تقـريـر  إجمالي بمعاملات الخدمات الداعمة
            FileName = "TRANSACTIONPAYMENTDEPENDANTSERVICESSUMMARY_sec"
            sheet = 'TRANSACTIONPAYMENTDEPENDANTSERV'
            latest_Supporting_Document_FileName = get_latest_excel_file(folder_path, FileName)

            net_Total_Supporting = get_last_value_in_column(latest_Supporting_Document_FileName, sheet,["C","D","E","F","G","H"])
            print_arabic(f"إجمالي القيمة المضافة في تقرير إجمالي الرسوم الضريبية: {net_Total_Supporting}")
            SaveToExcel(f"إجمالي القيمة المضافة في تقرير إجمالي الرسوم الضريبية",net_Total_Supporting,folder_path)

            rowWord = "سند قبض"
            receipt = get_valueOf_receipt_document(latest_Supporting_Document_FileName,rowWord)

            #############################################################
            FileName = "TRANSACTIONTAXSUMARY"
            sheet = 'TRANSACTIONTAXSUMARY'
            latest_file_VAT_From_Total_VAT_FileName = get_latest_excel_file(folder_path, FileName)

            TotalVAT_Report = get_last_value_in_column(latest_file_VAT_From_Total_VAT_FileName, sheet,["I","J"])
            print_arabic(f"إجمالي القيمة المضافة في تقرير إجمالي الرسوم الضريبية: {TotalVAT_Report}")
            SaveToExcel(f"إجمالي القيمة المضافة في تقرير إجمالي الرسوم الضريبية",TotalVAT_Report,folder_path)

            #############################################################              
            # أمانات
            FileName = "TransactionFeeReportforAllPaymentMethods(Deposit)"
            latest_Tahseel_Trusrt_File = get_latest_excel_file(folder_path, FileName)
            sheet = 'TransactionFeesForAllPaymentMet'

            last_Tahseel_Trust_Fee = get_last_value_in_column(latest_Tahseel_Trusrt_File, sheet,['AF','AG','AH','AI'])
            print_arabic(f"إجمالي رسوم الخدمات - تحصيل في تقرير كافة وسائل الدفع (أمانات) :  {last_Tahseel_Trust_Fee}")
            SaveToExcel(f"إجمالي رسوم الخدمات - تحصيل في تقرير كافة وسائل الدفع (أمانات)",last_Tahseel_Trust_Fee,folder_path)

            last_SFD_Trust_ServiceFee = get_last_value_in_column(latest_Tahseel_Trusrt_File, sheet,['AJ','AK','AL'])
            print_arabic(f"إجمالي رسوم الخدمات - المالية المركزية في تقرير كافة وسائل الدفع (أمانات): {last_SFD_Trust_ServiceFee}")
            SaveToExcel(f"إجمالي رسوم الخدمات - المالية المركزية في تقرير كافة وسائل الدفع (أمانات)",last_SFD_Trust_ServiceFee,folder_path)


            #ضريبة تحصيل أمانات
            Service_Vat_Tahseel_Share = get_last_value_in_column(latest_Tahseel_Trusrt_File, sheet,"AE")
            print_arabic(f"ضريبة تحصيل (أمانات) {Service_Vat_Tahseel_Share}")
            SaveToExcel(f"ضريبة تحصيل (أمانات)",Service_Vat_Tahseel_Share,folder_path)

            last_FeeValue_AllPaymentMethods_Deposit = get_last_value_in_column(latest_Tahseel_Trusrt_File, sheet,"AM")
            print_arabic(f"إجمالي رسوم المعاملة في تقرير كافة وسائل الدفع - (أمانات) {last_FeeValue_AllPaymentMethods_Deposit}")
            SaveToExcel(f"إجمالي رسوم المعاملة في تقرير كافة وسائل الدفع - (أمانات)",last_FeeValue_AllPaymentMethods_Deposit,folder_path)

            #############################################################
            #تقرير إجمالي التحملات حسب جهة الإيراد

            FileName = "SummaryReport_of_IncurredFees PerRevenueEntity"
            latest_Fees_File = get_latest_excel_file(folder_path, FileName)
            sheet = 'SummaryReport_of_IncurredFees P'

            last_ServiceFees = get_last_value_in_column(latest_Fees_File, sheet,"U")
            print_arabic(f"إجمالي الرسوم من تقرير إجمالي التحملات حسب جهة الإيراد {last_ServiceFees}")
            SaveToExcel(f"إجمالي الرسوم من تقرير إجمالي التحملات حسب جهة الإيراد",last_ServiceFees,folder_path)

            last_BackFees_TotalCharges = get_last_value_in_column(latest_Fees_File, sheet,"L")
            print_arabic(f"إجمالي الرسوم البنكية من تقرير إجمالي التحملات حسب جهة الإيراد {last_BackFees_TotalCharges}")
            SaveToExcel(f"إجمالي الرسوم البنكية من تقرير إجمالي التحملات حسب جهة الإيراد",last_BackFees_TotalCharges,folder_path)

            last_VAT_TotalCharges = get_last_value_in_column(latest_Fees_File, sheet,["R","S","T"])
            print_arabic(f"إجمالي ضريبة رسوم الخدمات من تقرير إجمالي التحملات حسب جهة الإيراد {last_VAT_TotalCharges}")
            SaveToExcel(f"إجمالي ضريبة رسوم الخدمات من تقرير إجمالي التحملات حسب جهة الإيراد",last_VAT_TotalCharges,folder_path)

            last_BankFeesVAT_TotalCharges = get_last_value_in_column(latest_Fees_File, sheet,"K")
            print_arabic(f"إجمالي ضريبة الرسوم البنكية من تقرير إجمالي التحملات حسب جهة الإيراد {last_BankFeesVAT_TotalCharges}")
            SaveToExcel(f"إجمالي ضريبة الرسوم البنكية من تقرير إجمالي التحملات حسب جهة الإيراد",last_BankFeesVAT_TotalCharges,folder_path)

            #############################################################
            # الإيصال الذكي

            FileName = "GITFees_ShjGovTransStatistics"
            latest_Smart_Receipt_File = get_latest_excel_file(folder_path, FileName)
            sheet = 'هيئة مطار الشارقة'
            Tahseel_Share_SmartReceipt = "رسوم الخدمات (درهم)"

            last_Tahseel_Share_SmartReceipt = get_last_value_in_column(latest_Smart_Receipt_File, sheet,"H")

            print_arabic(f"إجمالي رسوم الخدمات - حصة الشركة من تقرير الإيصال الذكي {last_Tahseel_Share_SmartReceipt}")
            SaveToExcel(f"إجمالي رسوم الخدمات - حصة الشركة من تقرير الإيصال الذكي",last_Tahseel_Share_SmartReceipt,folder_path)

            last_SFD_Share_SmartReceipt = get_last_value_in_column(latest_Smart_Receipt_File, sheet,["I","J","K"])
            print_arabic(f"إجمالي رسوم الخدمات - حصة المالية من تقرير الإيصال الذكي {last_SFD_Share_SmartReceipt}")
            SaveToExcel(f"إجمالي رسوم الخدمات - حصة المالية من تقرير الإيصال الذكي",last_SFD_Share_SmartReceipt,folder_path)

            Service_Vat_SFD_Share = get_last_value_in_column(latest_Smart_Receipt_File, sheet,["F","G"])
            print_arabic(f" إجمالي ضريبة رسوم الخدمات - حصة المالية في تقرير الإيصال الذكي {Service_Vat_SFD_Share}")
            SaveToExcel(f" إجمالي ضريبة رسوم الخدمات - حصة المالية في تقرير الإيصال الذكي",Service_Vat_SFD_Share,folder_path)

            Service_Vat_Tahseel_Share = get_last_value_in_column(latest_Smart_Receipt_File, sheet,"E")
            print_arabic(f"إجمالي ضريبة رسوم الخدمات - حصة الشركة في تقرير الإيصال الذكي {Service_Vat_SFD_Share}")
            SaveToExcel(f"إجمالي ضريبة رسوم الخدمات - حصة الشركة في تقرير الإيصال الذكي",Service_Vat_SFD_Share,folder_path)

            Total_VAT_From_SmartReceipt_Report = Service_Vat_SFD_Share + Service_Vat_Tahseel_Share

            # Receipt_voucher
            ################################################################################################################################
            #region Calculations Section
            # فرق قيمة رسوم المعاملة
            compare_values("إجمالي الرسوم في تقرير المقبوضات",last_transaction_value_Revenue,"إجمالي الرسوم في تقرير كافة وسائل الدفع",last_transaction_value)
            SaveToExcel("الفرق بين إجمالي الرسوم في تقرير المقبوضات وإجمالي الرسوم في تقرير كافة وسائل الدفع",last_transaction_value_Revenue-last_transaction_value,folder_path)
            # فرق قيمة رسوم المعاملة 1
            compare_values("إجمالي رسوم المعاملة من تقرير كافة وسائل الدفع",last_transaction_value,"إجمالي رسوم المعاملات من تقرير إجمالي المعاملات",Total_Transactions_all_dep)
            SaveToExcel("الفرق بين إجمالي رسوم المعاملة من تقرير كافة وسائل الدفع وإجمالي رسوم المعاملات من تقرير إجمالي المعاملات",last_transaction_value-Total_Transactions_all_dep,folder_path)

            # فرق ضريبية القيمة المضافة 1
            compare_values("إجمالي القيمة المضافة في تقرير المقبوضات",last_Tax_Total_Revenue,"إجمالي الضريبة في تقرير كافة وسائل الدفع",last_tax_value)
            SaveToExcel("الفرق بين إجمالي القيمة المضافة في تقرير المقبوضات وإجمالي الضريبة في تقرير كافة وسائل الدفع",last_Tax_Total_Revenue-last_tax_value,folder_path)

            # فرق ضريبية القيمة المضافة 2
            compare_values("إجمالي الضريبة من تقرير كافة وسائل الدفع",last_tax_value,"إجمالي الضريبة من تقرير إجمالي الرسوم الضريبية",TotalVAT_Report)
            SaveToExcel("الفرق بين إجمالي الضريبة من تقرير كافة وسائل الدفع وإجمالي الضريبة من تقرير إجمالي الرسوم الضريبية",last_tax_value-TotalVAT_Report,folder_path)

            # فرق رسوم الخدمات - المالية المركزية
            compare_values("إجمالي رسوم الخدمات - الماليةالمركزية (إيراد + أمانات)",last_SFD_ServiceFee+last_SFD_Trust_ServiceFee,"حصة المالية من تقرير الإيصال الذكي",last_SFD_Share_SmartReceipt)
            SaveToExcel("الفرق بين إجمالي رسوم الخدمات - المالية المركزية (إيراد+أمانات) وحصة المالية من تقرير الإيصال الذكي",(last_SFD_ServiceFee+last_SFD_Trust_ServiceFee)-last_SFD_Share_SmartReceipt,folder_path)

            # فرق رسوم الخدمات - تحصيل
            compare_values("إجمالي رسوم الخدمات - تحصيل (إيراد + أمانات)",last_Tahseel_Fee + last_Tahseel_Trust_Fee,"حصة الشركة من تقرير الإيصال الذكي",last_Tahseel_Share_SmartReceipt)
            SaveToExcel("الفرق بين إجمالي رسوم الخدمات - تحصيل (إيراد + أمانات) و حصة الشركة من تقرير الإيصال الذكي",(last_Tahseel_Fee + last_Tahseel_Trust_Fee)-last_Tahseel_Share_SmartReceipt,folder_path)

            # فرق رسوم الخدمات العالمية
            compare_values("إجمالي رسوم الخدمات في تقرير إجمالي معاملات بطاقة الائتمان",Uni_CreditRepprt_TotalServiceFee,"إجمالي رسوم خدمات المالية + رسوم خدمات تحصيل في تقرير كافة وسائل الدفع (وسائل الدفع : المحافظ العالمية)",Uni_TotalServiceFee)
            SaveToExcel("الفرق بين إجمالي رسوم خدمات المالية + رسوم خدمات تحصيل في تقرير كافة وسائل الدفع (المحافظ العالمية) - إجمالي رسوم الخدمات في تقرير معاملات بطاقات الائتمان",Uni_CreditRepprt_TotalServiceFee-Uni_TotalServiceFee,folder_path)

            # فرق ضريبة تحصيل
            compare_values("إجمالي ضريبة المخرجات في تقرير معاملات بطاقات الائتمان",Uni_CreditRepprt_OutputFee,"إجمالي ضريبة تحصيل في تقرير كافة وسائل الدفع (وسائل الدفع: المحافظ العالمية)",Uni_TahseelVAT)
            SaveToExcel("الفرق بين إجمالي ضريبة تحصيل في تقرير كافة وسائل الدفع (وسائل الدفع: المحافظ العالمية) - إجمالي ضريبة المخرجات في تقرير معاملات بطاقات الائتمان",Uni_CreditRepprt_OutputFee-Uni_TahseelVAT,folder_path)

            # فرق رسوم البنك
            compare_values("إجمالي رسوم البنك في تقرير معاملات بطاقات الائتمان",Uni_CreditRepprt_BankFee,"إجمالي المصاريف البنكية في تقرير كافة وسائل الدفع ( وسائل الدفع: المحافظ العالمية)+ إجمالي تحمل الرسوم البنكية في تقرير وسائل الدفع(المحافظ العالمية)",Uni_BankFeeAndBearExp)
            SaveToExcel("الفرق بين إجمالي المصاريف البنكية في تقرير كافة وسائل الدفع ( وسائل الدفع: المحافظ العالمية)+ إجمالي تحمل الرسوم البنكية في تقرير وسائل الدفع(المحافظ العالمية) - إجمالي رسوم البنك في تقرير معاملات بطاقات الائتمان",Uni_CreditRepprt_BankFee-Uni_BankFeeAndBearExp,folder_path)

            compare_values("إجمالي رسوم البنك ضريبة المخرجات في تقرير معاملات بطاقات الائتمان",Uni_CreditRepprt_OutputTax,"إجمالي ض.تحمل الرسوم البنكية + ضريبة المصاريف البنكية في تقرير كافة وسائل الدفع ( المحافظ العالمية)",Universal_BankVatFee)
            SaveToExcel("الفرق بين إجمالي ض.تحمل الرسوم البنكية + ضريبة المصاريف البنكية في تقرير كافة وسائل الدفع ( المحافظ العالمية) -  إجمالي رسوم البنك ضريبة المخرجات في تقرير معاملات بطاقات الائتمان",Universal_BankVatFee-Uni_CreditRepprt_OutputTax,folder_path)

            compare_values("إجمالي قيمة الرسوم في تقرير إجمالي بالمعاملات حسب جهة الأمانة - المقبوضات",last_FeeValue_Depositreceivable,"إجمالي رسوم المعاملة في تقرير كافة وسائل الدفع - (أمانات)",last_FeeValue_AllPaymentMethods_Deposit)
            SaveToExcel("إجمالي قيمة الرسوم في تقرير إجمالي بالمعاملات حسب جهة الأمانة - المقبوضات - إجمالي رسوم المعاملة في تقرير كافة وسائل الدفع - (أمانات)",last_FeeValue_Depositreceivable-last_FeeValue_AllPaymentMethods_Deposit,folder_path)


            compare_values("إجمالي صافي الرسوم في تقرير كافة وسائل الدفع (المحافظ العالمية)",bankPayments_NetFees,"إجمالي المالية في تقرير معاملات بطاقات الائتمان ",Uni_SFDTotal)
            SaveToExcel("الفرق بين إجمالي المالية في تقرير معاملات بطاقات الائتمان وإجمالي صافي الرسوم في تقرير كافة وسائل الدفع (المحافظ العالمية)",bankPayments_NetFees-Uni_SFDTotal,folder_path)

            compare_values("إجمالي ضريبة المدخلات في توزيع رسوم البنك في تقرير معاملات بطاقات الائتمان",Uni_DistributeBankFees_IncomeFee,"إجمالي ضريبة المدخلات في تقرير كافة وسائل الدفع (المحافظ العالمية) ",bankPayments_IncomeFees)
            SaveToExcel("الفرق بين إجمالي ضريبة المدخلات في تقرير كافة وسائل الدفع (المحافظ العالمية)وإجمالي ضريبة المدخلات في توزيع رسوم البنك في تقرير معاملات بطاقات الائتمان",Uni_DistributeBankFees_IncomeFee-bankPayments_IncomeFees,folder_path)

            compare_values("إجمالي حصة البنك في توزيع رسوم البنك في تقرير معاملات بطاقات الائتمان",Uni_DistributeBankFees_Bank,"إجمالي حصة البنك في تقرير كافة وسائل الدفع (المحافظ العالمية) ",bankPayments_BankShare)
            SaveToExcel("إجمالي حصة البنك في تقرير كافة وسائل الدفع (المحافظ العالمية)وإجمالي حصة البنك في توزيع رسوم البنك في تقرير معاملات بطاقات الائتمان",Uni_DistributeBankFees_Bank-bankPayments_BankShare,folder_path)

            compare_values("إجمالي تحمل الجهة في توزيع رسوم البنك في تقرير معاملات بطاقات الائتمان",Uni_DistributeBankFees_EntityHold,"إجمالي تحمل الرسوم البنكية + ضريبة تحمل الرسوم البنكية في تقرير كافة وسائل الدفع (المحافظ العالمية) ",Uni_BankFeePlusBankFeeVAT)
            SaveToExcel("الفرق بين إجمالي تحمل الرسوم البنكية + ضريبة تحمل الرسوم البنكية في تقرير كافة وسائل الدفع (المحافظ العالمية)و إجمالي تحمل الجهة في توزيع رسوم البنك في تقرير معاملات بطاقات الائتمان",Uni_DistributeBankFees_EntityHold-Uni_BankFeePlusBankFeeVAT,folder_path)
            
            compare_values("إجمالي رسوم البنك + ضريبة مخرجات رسوم البنك في تقرير معاملات بطاقات الائتمان",Uni_BankFeesPlusOutputTax,"إجمالي الرسوم البنكية في تقرير كافة وسائل الدفع (المحافظ العالمية) ",bankPayments_TotalBankFees)
            SaveToExcel("الفرق بين إجمالي الرسوم البنكية في تقرير كافة وسائل الدفع (المحافظ العالمية)و إجمالي رسوم البنك + ضريبة مخرجات رسوم البنك في تقرير معاملات بطاقات الائتمان",Uni_BankFeesPlusOutputTax-bankPayments_TotalBankFees,folder_path)
            
            compare_values("إجمالي رسوم المعاملة لإجمالي رسوم الجهة في تقرير وسائل الدفع (المحافظ العالمية)",Uni_TotalTransactionFeeForEntityFee,"إجمالي رسوم المعاملة في تقرير معاملات بطاقات الائتمان ",Uni_TotalTransactionFee)
            SaveToExcel("الفرق بين إجمالي رسوم المعاملة في تقرير معاملات بطاقات الائتمان و إجمالي رسوم المعاملة لإجمالي رسوم الجهة في تقرير وسائل الدفع (المحافظ العالمية)",Uni_TotalTransactionFeeForEntityFee-Uni_TotalTransactionFee,folder_path)
            
            compare_values("إجمالي الضريبة لإجمالي رسوم الجهة في تقرير وسائل الدفع (المحافظ العالمية)",Uni_TotalVATForEntityFee,"إجمالي ضريبة المخرجات لرسوم المعاملة في تقرير معاملات بطاقات الائتمان ",Uni_TotalTransactionFeeVat)
            SaveToExcel("الفرق بين إجمالي ضريبة المخرجات لرسوم المعاملة في تقرير معاملات بطاقات الائتمان و إجمالي الضريبة لإجمالي رسوم الجهة في تقرير وسائل الدفع (المحافظ العالمية)",Uni_TotalVATForEntityFee-Uni_TotalTransactionFeeVat,folder_path)
            
            compare_values("إجمالي رسوم دعم الأبحاث في تقرير معاملات بطاقات الائتمان",Uni_ResearchSupportFees,"إجمالي رسوم البحث العلمي في تقرير وسائل الدفع (المحافظ العالمية) ",Uni_ResarchFees)
            SaveToExcel("الفرق بين إجمالي رسوم البحث العلمي في تقرير وسائل الدفع (المحافظ العالمية) و إجمالي رسوم دعم الأبحاث في تقرير معاملات بطاقات الائتمان",Uni_ResearchSupportFees-Uni_ResarchFees,folder_path)
            
            compare_values("إجمالي تحمل الجهة لرسوم دعم الأبحاث في تقرير معاملات بطاقات الائتمان",Uni_EntityHold,"إجمالي تحمل الجهة للرسوم في تقرير وسائل الدفع (المحافظ العالمية) ",Uni_EntityShareFeesForResarchFees)
            SaveToExcel("الفرق بين إجمالي تحمل الجهة للرسوم في تقرير وسائل الدفع (المحافظ العالمية) و إجمالي تحمل الجهة لرسوم دعم الأبحاث في تقرير معاملات بطاقات الائتمان",Uni_EntityHold-Uni_EntityShareFeesForResarchFees,folder_path)
            

            TotalFeesCoverage = last_ServiceFee_Charges + last_BankFees + last_ServiceFees + last_BackFees_TotalCharges
            TotalFeesCoverageVAT = last_VAT_Charge +last_BankFees_VAT + last_VAT_TotalCharges +last_BankFeesVAT_TotalCharges
            print_arabic(f" إجمالي تحمل الرسوم {round(TotalFeesCoverage, 4)}")
            SaveToExcel(f"إجمالي تحمل الرسوم",round(TotalFeesCoverage, 4),folder_path)

            print_arabic(f"إجمالي ضريبة تحمل الرسوم {round(TotalFeesCoverageVAT,4)}")
            SaveToExcel(f"إجمالي ضريبة تحمل الرسوم",round(TotalFeesCoverageVAT,4),folder_path)


            compare_values("إجمالي تحمل الرسوم",round(TotalFeesCoverage, 4),"إجمالي ضريبة تحمل الرسوم",round(TotalFeesCoverageVAT,4),1)
            SaveToExcel(f"قيمة سند القبض في تقرير إجمالي إيراد الخدمات الداعمة ",round(TotalFeesCoverage, 4) + round(TotalFeesCoverageVAT,4),folder_path)
            
            print_arabic(f" قيمة سند القبض : {net_Total_Supporting-Resacrch_Fees}")
            SaveToExcel(f"قيمة سند القبض ",net_Total_Supporting-Resacrch_Fees,folder_path)

            print_arabic(f"قيمة سند القبض في تقرير إجمالي إيراد الخدمات الداعمة : {receipt}")
            SaveToExcel(f"قيمة سند القبض في تقرير إجمالي إيراد الخدمات الداعمة ",receipt,folder_path)

        if receipt is not None and not math.isnan(receipt):
            compare_values("قيمة سند القبض",net_Total_Supporting-Resacrch_Fees,"قيمة سند القبض في تقرير إجمالي إيراد الخدمات الداعمة ",receipt)
            SaveToExcel(f"الفرق بين قيمة سند القبض  وقيمة سند القبض في تقرير إجمالي إيراد الخدمات الداعمة",(net_Total_Supporting-Resacrch_Fees)-receipt,folder_path)

    except Exception as e:
        print_arabic("❌ حدثت المشكلة التالية", type(e).__name__, "-", str(e))

    finally:
        input("⏸️ Press Enter to close the browser...")
        # driver.quit()